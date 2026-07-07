from pathlib import Path

import pytest
from openpyxl import load_workbook

from excel_builder.config import MasterSourcesReader
from excel_builder.guards import ImmutableMasterGuard, ImmutableMasterViolation
from excel_builder.io import WorkbookWriter
from excel_builder.models import BuilderInputRow, BuilderResult


def test_master_sources_config_points_to_v1_files():
    sources = MasterSourcesReader().read()

    assert sources.master_version == "v1.0"
    assert sources.word_master == Path("data/word_templates/Taxestruktur_Master_v1.0.docx")
    assert sources.excel_master == Path("data/master_templates/ArbetsExcel_Template_v1.0.xlsx")
    assert sources.immutable is True
    assert sources.word_master.exists()
    assert sources.excel_master.exists()


def test_immutable_guard_refuses_master_output():
    sources = MasterSourcesReader().read()
    guard = ImmutableMasterGuard([sources.excel_master])

    with pytest.raises(ImmutableMasterViolation):
        guard.assert_output_allowed(sources.excel_master)


def test_immutable_guard_protects_taxepunkter_a_to_e_and_taxa_fran_edp():
    guard = ImmutableMasterGuard([])

    with pytest.raises(ImmutableMasterViolation):
        guard.assert_sheet_write_allowed("Taxepunkter", 5)

    guard.assert_sheet_write_allowed("Taxepunkter", 6)

    with pytest.raises(ImmutableMasterViolation):
        guard.assert_sheet_write_allowed("Taxa_från_edp", 10)


def test_workbook_writer_creates_copy_without_changing_master(tmp_path):
    sources = MasterSourcesReader().read()
    guard = ImmutableMasterGuard([sources.excel_master])
    before = guard.fingerprint(sources.excel_master)

    out = tmp_path / "ArbetsExcel_copy.xlsx"
    result = BuilderResult(rows=[
        BuilderInputRow(section="2.1", name="En- och tvåbostadshus", unit="st/år"),
    ])

    WorkbookWriter().write(result, out)

    assert out.exists()
    guard.verify_unchanged(before)
    wb = load_workbook(out, read_only=True, data_only=False)
    assert "Taxepunkter" in wb.sheetnames
    assert "Taxa_från_edp" in wb.sheetnames
    assert "Builder_Output" in wb.sheetnames
    assert wb["Builder_Output"]["C2"].value == "En- och tvåbostadshus"
