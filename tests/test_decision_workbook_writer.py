from openpyxl import Workbook, load_workbook

from excel_builder.decision import DecisionWorkbookWriter
from excel_builder.edp import ProposalTraceSheets
from excel_builder.models import ParserTaxRow, TaxDecision, TaxDecisionReport


def test_decision_workbook_writer_adds_status_columns_and_trace(tmp_path):
    workbook = tmp_path / "arbets.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Taxepunkter"
    ws.append(["Paragraf", "Taxapunkt"])
    ws.append(["6.1.2", "Asbest, emballerat"])
    ProposalTraceSheets().add(wb, municipality="Sorsele", context="test")
    wb.save(workbook)

    report = TaxDecisionReport(decisions=[
        TaxDecision(
            parser_row=ParserTaxRow(section="6.1.2", tax_point="Asbest, emballerat"),
            status="NEW_TAXA",
            source="Word",
            rule="testregel",
            confidence=0.0,
            comment="testkommentar",
        )
    ])

    DecisionWorkbookWriter().write(workbook, report, municipality="Sorsele")

    result = load_workbook(workbook)
    ws = result["Taxepunkter"]

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "Beslutsstatus" in headers
    assert "Beslutsregel" in headers
    assert "Beslutskommentar" in headers
    assert "NEW_TAXA" in [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]

    trace = result["Regelspårning"]
    assert trace["D7"].value == "Beslut"
    assert trace["E7"].value == "NEW_TAXA"
