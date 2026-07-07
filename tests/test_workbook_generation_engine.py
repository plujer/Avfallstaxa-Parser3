from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_builder.models import WorkbookDecisionRow
from excel_builder.workbook_generation import WorkbookGenerationEngine, DecisionTraceCsvReader


def test_workbook_generation_adds_decision_trace_and_taxepunkter_columns(tmp_path: Path):
    workbook = tmp_path / "arbets.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Taxepunkter"
    ws.append(["Taxapunkt", "EDP taxekod"])
    ws.append(["Hämtning 240 liter rest", ""])
    wb.save(workbook)

    rows = [
        WorkbookDecisionRow(
            word_tax_code="KÄ240RE26",
            candidate_tax_code="KÄ240RE26",
            decision="ACCEPT",
            confidence=0.94,
            total_score=0.91,
            primary_reason="Samma taxefamilj och variant.",
            signals="family:1.00*0.30=0.3000",
        )
    ]

    report = WorkbookGenerationEngine().write(workbook, rows)

    assert report.rows_written == 1
    assert report.taxepunkter_rows_updated == 1
    result = load_workbook(workbook)
    assert "Decision_Trace" in result.sheetnames
    assert "Workbook_Generation" in result.sheetnames
    tax_ws = result["Taxepunkter"]
    headers = [cell.value for cell in tax_ws[1]]
    assert "Beslutsspår kandidat" in headers
    assert "Beslutsspår status" in headers
    assert tax_ws.cell(2, headers.index("Beslutsspår status") + 1).value == "ACCEPT"


def test_decision_trace_reader_reads_semicolon_csv(tmp_path: Path):
    source = tmp_path / "decision_traces.csv"
    source.write_text(
        "word_tax_code;candidate_tax_code;decision;confidence;total_score;primary_reason;rejected_reason;signals\n"
        "KÄ140MA26;KÄ140MA26;ACCEPT;0,9500;0,9200;Match; ;family:1\n",
        encoding="utf-8-sig",
    )

    rows = DecisionTraceCsvReader().read(source)

    assert len(rows) == 1
    assert rows[0].word_tax_code == "KÄ140MA26"
    assert rows[0].confidence == 0.95
