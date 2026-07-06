from openpyxl import Workbook, load_workbook

from excel_builder.edp import ProposalTraceSheets, IsolatedWorkbookBuilder, EdpExportReader
from excel_builder.io import WorkbookWriter
from excel_builder.models import BuilderInputRow, BuilderResult


def test_proposal_trace_sheets_are_added_to_workbook():
    wb = Workbook()
    wb.active.title = "Taxepunkter"

    ProposalTraceSheets().add(wb, municipality="Sorsele", context="test")

    assert "Taxa_Förslag" in wb.sheetnames
    assert "Regelspårning" in wb.sheetnames
    assert wb["Taxa_Förslag"]["A6"].value == "Kommun"
    assert wb["Regelspårning"]["A6"].value == "Kommun"


def test_isolated_workbook_builder_includes_proposal_and_trace_sheets(tmp_path):
    edp = tmp_path / "edp.xlsx"

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Rapport"])
    ws.append(IsolatedWorkbookBuilder.EDP_HEADERS)
    ws.append([1, "KOD1", "Test", "RENH", "HUSH", 1, "ÅRPR", "Taxedel", "", "", 100, 45658, 0, "", "-", 1, "FORMEL"])
    wb.save(edp)

    export = EdpExportReader().read(edp, "Sorsele")
    out = tmp_path / "ArbetsExcel_Sorsele_byggd.xlsx"

    IsolatedWorkbookBuilder(include_standard_tax_sheets=False).build(export, out)

    result = load_workbook(out)
    assert "Taxa_Förslag" in result.sheetnames
    assert "Regelspårning" in result.sheetnames
    assert result["Taxa_Förslag"]["B3"].value == "Sorsele"
    assert result["Regelspårning"]["B3"].value == "Sorsele"


def test_parser_workbook_writer_includes_proposal_and_trace_sheets(tmp_path):
    out = tmp_path / "arbets.xlsx"
    result = BuilderResult(rows=[
        BuilderInputRow(section="6.1.2", name="Asbest, emballerat", unit="kilogram")
    ])

    WorkbookWriter().write(result, out)

    wb = load_workbook(out)
    assert "Taxepunkter" in wb.sheetnames
    assert "Taxa_Förslag" in wb.sheetnames
    assert "Regelspårning" in wb.sheetnames
