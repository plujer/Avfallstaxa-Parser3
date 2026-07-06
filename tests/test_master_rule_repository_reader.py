from openpyxl import Workbook

from excel_builder.rules import MasterRuleRepositoryReader


def test_master_rule_repository_reader_reads_taxepunkter_and_edp(tmp_path):
    path = tmp_path / "master.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Taxepunkter"
    ws.append(["Paragraf", "Taxapunkt", "Enhet", "Taxakod", "Formel", "Taxedel"])
    ws.append(["6.1.2", "Asbest, emballerat", "kilogram", "EDP1", "FORMEL1", "Taxedel"])

    edp = wb.create_sheet("Taxa_från_edp")
    edp.append(["meta"])
    edp.append(["strTaxekod", "strTaxebenamning", "strFaktor", "strTaxedelAvser", "strFormel"])
    edp.append(["EDP2", "Gips", "VIKG", "Kilogram", "FORMEL2"])

    doc = wb.create_sheet("Dokumentation_Taxepunkter")
    doc.append(["Regel", "Text"])
    doc.append(["Taxakod", "Taxakod får inte ändras automatiskt"])

    wb.save(path)

    repo = MasterRuleRepositoryReader().read(path)

    assert repo.rule_count >= 3
    assert len(repo.taxepunkt_rules) == 1
    assert len(repo.edp_rules) == 1
    assert len(repo.documentation_rules) >= 1
    assert repo.taxepunkt_rules[0].tax_code == "EDP1"


def test_master_rule_repository_reader_warns_missing_workbook(tmp_path):
    repo = MasterRuleRepositoryReader().read(tmp_path / "missing.xlsx")

    assert repo.rule_count == 0
    assert repo.warnings
