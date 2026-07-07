from pathlib import Path
from openpyxl import Workbook

from excel_builder.template import TemplateMasterManager


def test_template_manager_creates_working_copy(tmp_path):
    template = tmp_path / "ArbetsExcel_Template_v0.1.0_draft.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Taxepunkter"
    ws["A1"] = "Mall"
    wb.save(template)

    out = tmp_path / "output" / "ArbetsExcel_Sorsele.xlsx"

    info = TemplateMasterManager().create_working_copy(out, template)

    assert out.exists()
    assert info.version == "v0.1.0"
    assert not info.warnings
    assert template.exists()


def test_template_manager_refuses_overwrite_template(tmp_path):
    template = tmp_path / "ArbetsExcel_Template_v0.1.0_draft.xlsx"
    wb = Workbook()
    wb.save(template)

    info = TemplateMasterManager().create_working_copy(template, template)

    assert info.warnings
    assert "Output får inte vara masterfil" in info.warnings[0]


def test_template_manager_proposes_versioned_template_name():
    name = TemplateMasterManager().propose_new_template_name("0.2.0", "draft")

    assert name == "ArbetsExcel_Template_v0.2.0_draft.xlsx"
