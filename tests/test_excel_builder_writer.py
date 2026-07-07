from openpyxl import load_workbook
from excel_builder.io import WorkbookWriter
from excel_builder.models import BuilderInputRow, BuilderResult


def test_workbook_writer_creates_arbets_excel(tmp_path):
    out = tmp_path / "arbets.xlsx"
    result = BuilderResult(rows=[
        BuilderInputRow(section="6.1.2", name="Asbest, emballerat", unit="kilogram")
    ])

    WorkbookWriter().write(result, out)

    wb = load_workbook(out)
    assert "Taxepunkter" in wb.sheetnames
    assert "Taxa_från_edp" in wb.sheetnames
    assert "Builder_Output" in wb.sheetnames
    assert "Sammanfattning" in wb.sheetnames
    assert "README" in wb.sheetnames
    ws = wb["Builder_Output"]
    assert ws["A1"].value == "Paragraf"
    assert ws["C2"].value == "Asbest, emballerat"
    assert ws["H2"].value == "Ej kopplad"
