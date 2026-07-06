from openpyxl import Workbook

from excel_builder.rules import RulebookReader, EdpRuleValidator


def create_rulebook_test_workbook(path):
    wb = Workbook()

    ws = wb.active
    ws.title = "03_Arbetsflöde"
    ws.append([
        "7",
        "Matcha mot Taxa_från_edp",
        "Använd matchningsordningen strTaxekod → strTaxebenamning → strTaxedelAvser → strFaktor → strFormel.",
    ])

    ws = wb.create_sheet("Dokumentation_Taxepunkter")
    ws.append([
        "F",
        "Taxakod",
        "Taxa_från_edp[strTaxekod]",
        "Taxakod får bara vara bekräftad EDP-kod.",
    ])
    ws.append([
        "J",
        "Aktuellt pris",
        "Taxa_från_edp[curNuvarandePris]",
        "Aktuellt pris får aldrig redigeras manuellt.",
    ])

    ws = wb.create_sheet("Dokumentation_Taxa_Saknas")
    ws.append([
        "Taxa_Saknas",
        "Saknade taxor",
        "Taxor som saknar EDP-matchning ska kunna läggas i Taxa_Saknas.",
    ])

    wb.save(path)


def test_rulebook_reader_reads_rule_sheets(tmp_path):
    path = tmp_path / "arbets.xlsx"
    create_rulebook_test_workbook(path)

    rulebook = RulebookReader().read(path)

    assert rulebook.count >= 4
    assert rulebook.contains_text("strTaxekod")
    assert rulebook.contains_text("strTaxebenamning")
    assert rulebook.contains_text("strTaxedelAvser")
    assert rulebook.contains_text("strFaktor")
    assert rulebook.contains_text("strFormel")
    assert rulebook.contains_text("Taxa_från_edp")
    assert rulebook.contains_text("Taxa_Saknas")
    assert rulebook.contains_text("Aktuellt pris")


def test_rulebook_reader_reports_missing_file():
    rulebook = RulebookReader().read("does_not_exist.xlsx")

    assert rulebook.count == 0
    assert rulebook.warnings
    assert "Arbets-Excel saknas" in rulebook.warnings[0]


def test_edp_rule_validator_detects_missing_terms():
    rulebook = RulebookReader().read("does_not_exist.xlsx")
    warnings = EdpRuleValidator().validate(rulebook)

    assert warnings


def test_edp_rule_validator_accepts_minimum_documented_terms(tmp_path):
    path = tmp_path / "arbets.xlsx"
    create_rulebook_test_workbook(path)

    rulebook = RulebookReader().read(path)
    warnings = EdpRuleValidator().validate(rulebook)

    assert warnings == []
