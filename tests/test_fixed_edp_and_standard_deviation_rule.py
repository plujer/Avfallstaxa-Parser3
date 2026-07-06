from openpyxl import Workbook, load_workbook

from excel_builder.edp import ProposalTraceSheets, IsolatedWorkbookBuilder, EdpExportReader


def test_proposal_trace_adds_edp_standard_deviation_sheet():
    wb = Workbook()
    wb.active.title = "Taxa_från_edp"

    ProposalTraceSheets().add(wb, municipality="Sorsele", context="test")

    assert "Taxa_Förslag" in wb.sheetnames
    assert "EDP_Avviker_Standard" in wb.sheetnames
    assert "Regelspårning" in wb.sheetnames
    assert wb["EDP_Avviker_Standard"]["A6"].value == "Kommun"
    assert "ändras aldrig automatiskt" in wb["EDP_Avviker_Standard"]["B2"].value


def test_isolated_workbook_documents_fixed_edp_rule(tmp_path):
    edp = tmp_path / "edp.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.append(["Rapport"])
    ws.append(IsolatedWorkbookBuilder.EDP_HEADERS)
    ws.append([1, "KOD1", "Test", "RENH", "HUSH", 1, "ÅRPR", "Taxedel", "", "", 100, 45658, 0, "", "-", 1, "FORMEL"])
    wb.save(edp)

    export = EdpExportReader().read(edp, "Sorsele")
    out = tmp_path / "ArbetsExcel_Sorsele_byggd.xlsx"

    IsolatedWorkbookBuilder(include_standard_tax_sheets=False).build(export, out)

    result = load_workbook(out)
    assert "EDP_Avviker_Standard" in result.sheetnames
    info_values = [result["Körningsinfo"].cell(row, 2).value for row in range(1, result["Körningsinfo"].max_row + 1)]
    assert any("Taxor i Taxa_från_edp är fasta" in str(value) for value in info_values)
