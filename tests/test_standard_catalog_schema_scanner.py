from openpyxl import Workbook, load_workbook

from excel_builder.standard import StandardCatalogSchemaScanner, StandardTaxReader, StandardCatalogNormalizer


def create_standard(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Standard Avfall"
    ws.append(["Intro"])
    ws.append(["strTaxekod", "strTaxebenamning", "strFaktor", "strTaxedelAvser", "strFormel"])
    ws.append(["STD1", "Asbest, emballerat", "VIKG", "Kilogram", "FORMEL1"])
    ws.append(["STD2", "Gips", "VIKG", "Kilogram", "FORMEL2"])
    ws.append([])
    ws.append([])
    ws.append(["taxekod", "benämning", "faktor"])
    ws.append(["STD3", "Träavfall", "VIKG"])
    wb.save(path)


def test_standard_catalog_schema_scanner_detects_multiple_sections(tmp_path):
    path = tmp_path / "standard.xlsx"
    create_standard(path)

    schema = StandardCatalogSchemaScanner().scan(path)

    assert schema.sheet_count == 1
    assert schema.section_count == 2
    assert schema.estimated_standard_rows == 3


def test_standard_tax_reader_reads_multiple_sections(tmp_path):
    path = tmp_path / "standard.xlsx"
    create_standard(path)

    catalog = StandardTaxReader().read(path)

    assert catalog.row_count == 3
    assert {row.strTaxekod for row in catalog.rows} == {"STD1", "STD2", "STD3"}


def test_standard_catalog_normalizer_writes_workbook(tmp_path):
    path = tmp_path / "standard.xlsx"
    create_standard(path)
    catalog = StandardTaxReader().read(path)

    out = tmp_path / "normalized.xlsx"
    StandardCatalogNormalizer().write(catalog, out)

    wb = load_workbook(out)
    assert "Standardtaxor_Normalized" in wb.sheetnames
    assert wb["Standardtaxor_Normalized"].max_row == 4
