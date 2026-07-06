from openpyxl import load_workbook

from excel_builder.edp import EdpExportReader, IsolatedWorkbookBuilder


def test_isolated_workbook_builder_creates_taxa_fran_edp_sheet(tmp_path):
    source = tmp_path / "edp.xlsx"

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Rapport"])
    ws.append(IsolatedWorkbookBuilder.EDP_HEADERS)
    ws.append([1, "KOD1", "Test", "RENH", "HUSH", 1, "ÅRPR", "Taxedel", "", "", 100, 45658, 0, "", "-", 1, "FORMEL"])
    wb.save(source)

    export = EdpExportReader().read(source, "Sorsele")
    out = tmp_path / "Sorsele" / "ArbetsExcel_Sorsele_byggd.xlsx"

    IsolatedWorkbookBuilder().build(export, out)

    result = load_workbook(out)
    assert "Taxa_från_edp" in result.sheetnames
    assert "Körningsinfo" in result.sheetnames
    assert result["Taxa_från_edp"]["B2"].value == "KOD1"
    assert result["Körningsinfo"]["B2"].value == "Sorsele"
