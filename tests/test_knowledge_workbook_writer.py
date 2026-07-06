from openpyxl import Workbook, load_workbook

from excel_builder.knowledge import KnowledgeWorkbookWriter, TaxKnowledgeExtractor
from excel_builder.models import ParserTaxRow


def test_knowledge_workbook_writer_adds_tax_knowledge_sheet(tmp_path):
    workbook = tmp_path / "arbets.xlsx"
    wb = Workbook()
    wb.active.title = "Taxepunkter"
    wb.save(workbook)

    rows = [ParserTaxRow(section="6.1.2", tax_point="Asbest, emballerat", unit="kilogram")]
    report = TaxKnowledgeExtractor().extract(rows)

    KnowledgeWorkbookWriter().write(workbook, report)

    result = load_workbook(workbook)
    assert "Tax_Knowledge" in result.sheetnames
    ws = result["Tax_Knowledge"]
    assert ws["A5"].value == "Paragraf"
    assert ws["B6"].value == "Asbest, emballerat"
    assert ws["F6"].value == "ÅVC/verksamhetsavfall"
    assert ws["J6"].value == "VIKG"


def test_knowledge_package_imports_writer():
    from excel_builder.knowledge import KnowledgeWorkbookWriter as ImportedWriter

    assert ImportedWriter is KnowledgeWorkbookWriter
