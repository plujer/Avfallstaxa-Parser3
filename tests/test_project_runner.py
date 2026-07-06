import json
from pathlib import Path
from openpyxl import Workbook, load_workbook

from excel_builder.models import ProjectConfig
from excel_builder.project import ProjectRunner
from excel_builder.edp import IsolatedWorkbookBuilder


def create_edp(path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Rapport"])
    ws.append(IsolatedWorkbookBuilder.EDP_HEADERS)
    ws.append([1, "KOD1", "Test", "RENH", "HUSH", 1, "ÅRPR", "Taxedel", "", "", 100, 45658, 0, "", "-", 1, "FORMEL"])
    wb.save(path)


def test_project_runner_creates_isolated_output(tmp_path):
    edp = tmp_path / "Sorsele.xlsx"
    create_edp(edp)

    config = ProjectConfig(
        municipality="Sorsele",
        word_path=str(tmp_path / "taxa.docx"),
        edp_export_path=str(edp),
        output_dir=str(tmp_path / "output" / "Sorsele"),
    )

    result = ProjectRunner().run(config)

    assert result.excel_path.endswith("ArbetsExcel_Sorsele_byggd.xlsx")
    assert Path(result.excel_path).exists()
    assert Path(result.manifest_path).exists()

    wb = load_workbook(result.excel_path)
    assert wb["Körningsinfo"]["B2"].value == "Sorsele"

    manifest = json.loads(Path(result.manifest_path).read_text(encoding="utf-8"))
    assert manifest["municipality"] == "Sorsele"
    assert "blandas" in manifest["isolation_rule"]
