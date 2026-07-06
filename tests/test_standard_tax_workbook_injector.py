from openpyxl import Workbook, load_workbook

from excel_builder.edp import StandardTaxWorkbookInjector, IsolatedWorkbookBuilder, EdpExportReader


def test_standard_tax_workbook_injector_adds_reference_sheet(tmp_path):
    standard = tmp_path / "standard.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Standardtaxor"
    ws.append(["strTaxekod", "strTaxebenamning"])
    ws.append(["STD1", "Standardtaxa"])
    wb.save(standard)

    target = Workbook()
    target.active.title = "Taxa_från_edp"

    warnings = StandardTaxWorkbookInjector().attach(target, standard)

    assert warnings == []
    assert any(name.startswith("EDP_Standard_") for name in target.sheetnames)


def test_isolated_workbook_builder_includes_standard_tax_reference(tmp_path):
    standard = tmp_path / "standard.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Standardtaxor"
    ws.append(["strTaxekod", "strTaxebenamning"])
    ws.append(["STD1", "Standardtaxa"])
    wb.save(standard)

    # Patch default standard path only inside this test.
    StandardTaxWorkbookInjector.DEFAULT_PATH = standard

    edp = tmp_path / "edp.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Rapport"])
    ws.append(IsolatedWorkbookBuilder.EDP_HEADERS)
    ws.append([1, "KOD1", "Test", "RENH", "HUSH", 1, "ÅRPR", "Taxedel", "", "", 100, 45658, 0, "", "-", 1, "FORMEL"])
    wb.save(edp)

    export = EdpExportReader().read(edp, "Sorsele")
    out = tmp_path / "ArbetsExcel_Sorsele_byggd.xlsx"

    IsolatedWorkbookBuilder().build(export, out)

    result = load_workbook(out)
    assert "Taxa_från_edp" in result.sheetnames
    assert any(name.startswith("EDP_Standard_") for name in result.sheetnames)
