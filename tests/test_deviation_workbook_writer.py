from openpyxl import Workbook, load_workbook

from excel_builder.edp import ProposalTraceSheets
from excel_builder.models import EdpExportRow, StandardTaxRow
from excel_builder.standard import EdpStandardDeviation, EdpStandardDeviationReport, DeviationWorkbookWriter


def test_deviation_workbook_writer_writes_review_rows(tmp_path):
    workbook = tmp_path / "arbets.xlsx"

    wb = Workbook()
    wb.active.title = "Taxa_från_edp"
    ProposalTraceSheets().add(wb, municipality="Sorsele", context="test")
    wb.save(workbook)

    report = EdpStandardDeviationReport(municipality="Sorsele", deviations=[
        EdpStandardDeviation(
            municipality="Sorsele",
            edp_row=EdpExportRow(strTaxekod="KOD1", strTaxebenamning="Kommun taxa", strFaktor="ANNAN"),
            standard_row=StandardTaxRow(source_sheet="Standard", row_number=2, strTaxekod="KOD1", strTaxebenamning="Kommun taxa", strFaktor="ÅRPR"),
            status="REVIEW",
            score=1.0,
            deviation_type="Faktor avviker",
            recommendation="Granska manuellt",
        )
    ])

    DeviationWorkbookWriter().write(workbook, report)

    result = load_workbook(workbook)
    ws = result["EDP_Avviker_Standard"]
    assert ws.max_row == 7
    assert ws["B7"].value == "KOD1"
    assert ws["J7"].value == "Faktor avviker"
    trace = result["Regelspårning"]
    assert trace["D7"].value == "EDP_Avviker_Standard"
