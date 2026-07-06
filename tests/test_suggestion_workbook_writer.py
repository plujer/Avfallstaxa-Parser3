from openpyxl import Workbook, load_workbook

from excel_builder.edp import ProposalTraceSheets
from excel_builder.models import ParserTaxRow, StandardTaxRow, StandardTaxSuggestion, StandardTaxSuggestionReport
from excel_builder.standard import SuggestionWorkbookWriter


def test_suggestion_workbook_writer_writes_proposals_and_trace(tmp_path):
    workbook = tmp_path / "arbets.xlsx"

    wb = Workbook()
    wb.active.title = "Taxepunkter"
    ProposalTraceSheets().add(wb, municipality="Sorsele", context="test")
    wb.save(workbook)

    report = StandardTaxSuggestionReport(suggestions=[
        StandardTaxSuggestion(
            parser_row=ParserTaxRow(section="6.1.2", tax_point="Asbest, emballerat", unit="kilogram"),
            standard_row=StandardTaxRow(
                source_sheet="Standard Avfall",
                row_number=2,
                strTaxekod="STD1",
                strTaxebenamning="Asbest, emballerat",
            ),
            status="PROPOSAL",
            score=0.95,
            method="test",
            comment="Stark träff",
        )
    ])

    SuggestionWorkbookWriter().write(workbook, report, municipality="Sorsele")

    result = load_workbook(workbook)
    forslag = result["Taxa_Förslag"]
    trace = result["Regelspårning"]

    assert forslag.max_row == 7
    assert forslag["A7"].value == "Sorsele"
    assert forslag["F7"].value == "STD1"
    assert trace["D7"].value == "Taxakod"
    assert trace["E7"].value == "STD1"


def test_suggestion_workbook_writer_does_not_write_no_suggestion(tmp_path):
    workbook = tmp_path / "arbets.xlsx"

    wb = Workbook()
    wb.active.title = "Taxepunkter"
    ProposalTraceSheets().add(wb, municipality="Sorsele", context="test")
    wb.save(workbook)

    report = StandardTaxSuggestionReport(suggestions=[
        StandardTaxSuggestion(
            parser_row=ParserTaxRow(section="6.1.2", tax_point="Okänd taxa"),
            standard_row=None,
            status="NO_SUGGESTION",
            score=0.0,
            method="test",
        )
    ])

    SuggestionWorkbookWriter().write(workbook, report, municipality="Sorsele")

    result = load_workbook(workbook)
    assert result["Taxa_Förslag"].max_row == 6
    assert result["Regelspårning"].max_row == 6
