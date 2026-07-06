"""Analyze a workbook and find likely facit sheets/columns."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from parser3.excel.excel_models import SheetProfile, WorkbookProfile


class WorkbookProfiler:
    FIELD_ALIASES = {
        "section": {
            "section", "paragraf", "kapitel", "§", "avsnitt", "taxeavsnitt",
            "taxestruktur", "struktur", "rubriknr"
        },
        "name": {
            "name", "namn", "taxepunkt", "taxa", "benämning", "benamning",
            "tjänst", "tjanst", "typ av avfall", "artikel", "produkt",
            "beskrivning", "avgift", "rubrik", "text"
        },
        "variant": {
            "variant", "intervall", "avgiftstyp", "tömningsintervall",
            "tomningsintervall", "frekvens", "hämtningsintervall",
            "hamtningsintervall"
        },
        "unit": {
            "unit", "enhet", "debiteringsenhet", "mängdenhet", "mangdenhet",
            "pris per", "per"
        },
        "price": {
            "price", "pris", "belopp", "avgift", "taxa", "kr", "kronor",
            "pris inkl moms", "pris inklusive moms"
        },
        "edp_code": {
            "edp", "taxekod", "kod", "edp kod", "edp-kod", "taxa kod",
            "artikelnummer", "tjänstekod", "tjanstekod"
        },
    }

    def profile(self, path: str | Path) -> WorkbookProfile:
        workbook_path = Path(path)
        wb = load_workbook(workbook_path, data_only=True)
        profile = WorkbookProfile(path=str(workbook_path))

        for ws in wb.worksheets:
            sheet_profile = SheetProfile(
                sheet_name=ws.title,
                max_row=ws.max_row,
                max_column=ws.max_column,
            )
            header_row, headers = self._find_header_row(ws)
            sheet_profile.header_row = header_row
            sheet_profile.headers = headers
            if header_row is not None:
                sheet_profile.detected_columns = self._detect_columns(headers)
                sheet_profile.candidate_score = self._score_sheet(sheet_profile, ws)
            profile.sheets.append(sheet_profile)

        return profile

    def _find_header_row(self, ws) -> tuple[int | None, list[str]]:
        best_row = None
        best_headers: list[str] = []
        best_score = 0

        for row_idx in range(1, min(ws.max_row, 40) + 1):
            values = [self._norm(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
            if not any(values):
                continue
            detected = self._detect_columns(values)
            score = len(detected) * 10
            if "name" in detected:
                score += 20
            if "price" in detected:
                score += 20
            if "section" in detected:
                score += 10
            if score > best_score:
                best_score = score
                best_row = row_idx
                best_headers = values

        return best_row, best_headers

    def _detect_columns(self, headers: list[str]) -> dict[str, int]:
        detected: dict[str, int] = {}
        for idx, header in enumerate(headers):
            clean = self._norm(header)
            if not clean:
                continue
            for field, aliases in self.FIELD_ALIASES.items():
                if field in detected:
                    continue
                if clean in aliases or any(alias in clean for alias in aliases if len(alias) >= 4):
                    detected[field] = idx + 1
        return detected

    def _score_sheet(self, profile: SheetProfile, ws) -> int:
        score = profile.candidate_score
        cols = profile.detected_columns
        if "name" in cols:
            score += 50
        if "price" in cols:
            score += 40
        if "section" in cols:
            score += 20
        if "unit" in cols:
            score += 10
        if "variant" in cols:
            score += 10

        # Add data presence score.
        if profile.header_row and "name" in cols:
            name_col = cols["name"]
            price_col = cols.get("price")
            data_rows = 0
            priced_rows = 0
            for row_idx in range(profile.header_row + 1, min(ws.max_row, profile.header_row + 200) + 1):
                name = self._norm(ws.cell(row_idx, name_col).value)
                if name:
                    data_rows += 1
                    if price_col and self._norm(ws.cell(row_idx, price_col).value):
                        priced_rows += 1
            score += min(data_rows, 50)
            score += min(priced_rows, 50)
        return score

    def _norm(self, value) -> str:
        return " ".join(str(value or "").replace("\xa0", " ").strip().lower().split())
