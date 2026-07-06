"""Read facit/master rows from the real Master.xlsx.

This reader first profiles all sheets and then reads the best candidate sheet.
It is deliberately tolerant because Master.xlsx can have Swedish column names,
merged header rows, and extra metadata columns.
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from parser3.excel.workbook_profiler import WorkbookProfiler
from parser3.models import TaxRow


class MasterExcelReader:
    def __init__(self) -> None:
        self.profiler = WorkbookProfiler()

    def read(self, path: str | Path, sheet_name: str | None = None) -> list[TaxRow]:
        workbook_path = Path(path)
        wb = load_workbook(workbook_path, data_only=True)
        profile = self.profiler.profile(workbook_path)

        sheet_profile = None
        if sheet_name:
            for item in profile.sheets:
                if item.sheet_name == sheet_name:
                    sheet_profile = item
                    break
        else:
            sheet_profile = profile.best_sheet

        if sheet_profile is None or sheet_profile.header_row is None:
            return []

        ws = wb[sheet_profile.sheet_name]
        cols = sheet_profile.detected_columns
        if "name" not in cols:
            return []

        result: list[TaxRow] = []
        for row_idx in range(sheet_profile.header_row + 1, ws.max_row + 1):
            name = self._cell(ws, row_idx, cols.get("name"))
            if not name or self._is_noise_row(name):
                continue

            section = self._cell(ws, row_idx, cols.get("section"))
            variant = self._cell(ws, row_idx, cols.get("variant"))
            unit = self._cell(ws, row_idx, cols.get("unit"))
            price = self._cell(ws, row_idx, cols.get("price"))
            edp_code = self._cell(ws, row_idx, cols.get("edp_code"))

            result.append(
                TaxRow(
                    section=section,
                    name=name,
                    variant=variant,
                    unit=unit,
                    price=price,
                    export=True,
                    group=edp_code,  # temporary storage until dedicated EDP model exists
                )
            )

        return result

    def _cell(self, ws, row_idx: int, col_idx: int | None) -> str:
        if not col_idx:
            return ""
        value = ws.cell(row_idx, col_idx).value
        return "" if value is None else str(value).strip()

    def _is_noise_row(self, name: str) -> bool:
        lower = " ".join(name.lower().split())
        if lower in {"", "summa", "totalt", "total", "taxa", "taxepunkt"}:
            return True
        if lower.startswith("kommentar"):
            return True
        return False
