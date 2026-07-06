"""Read existing Taxepunkter rows from Arbets-Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from excel_builder.models import WorkbookTaxRow


class WorkbookTaxepunkterReader:
    """Read Taxepunkter sheet without modifying workbook."""

    DEFAULT_HEADER_ROW = 5

    def read(self, workbook_path: str | Path, sheet_name: str = "Taxepunkter") -> list[WorkbookTaxRow]:
        path = Path(workbook_path)
        if not path.exists():
            return []

        wb = load_workbook(path, data_only=False, read_only=False)
        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        header_row = self._find_header_row(ws) or self.DEFAULT_HEADER_ROW

        rows: list[WorkbookTaxRow] = []

        for row_idx in range(header_row + 1, ws.max_row + 1):
            section = self._cell(ws, row_idx, 1)
            paragraph_name = self._cell(ws, row_idx, 2)
            tax_point = self._cell(ws, row_idx, 3)
            variant = self._cell(ws, row_idx, 4)
            unit = self._cell(ws, row_idx, 5)
            tax_code = self._cell(ws, row_idx, 6)
            proposed_price = self._cell(ws, row_idx, 7)

            if not any([section, tax_point, variant, unit, tax_code, proposed_price]):
                continue

            # Skip visual group/header rows where A:E are merged section labels.
            if tax_point == "" and variant == "" and unit == "" and tax_code == "":
                continue

            rows.append(
                WorkbookTaxRow(
                    row_number=row_idx,
                    section=section,
                    paragraph_name=paragraph_name,
                    tax_point=tax_point,
                    variant=variant,
                    unit=unit,
                    tax_code=tax_code,
                    proposed_price=proposed_price,
                )
            )

        return rows

    def _find_header_row(self, ws) -> int | None:
        for row_idx in range(1, min(ws.max_row, 30) + 1):
            values = [self._cell(ws, row_idx, col_idx).lower() for col_idx in range(1, min(ws.max_column, 14) + 1)]
            if "paragraf" in values and "taxapunkt" in values:
                return row_idx
        return None

    def _cell(self, ws, row: int, col: int) -> str:
        value = ws.cell(row, col).value
        if value is None:
            return ""
        if isinstance(value, str) and value.startswith("="):
            return value
        return str(value).strip()
