"""Create a compact workbook snapshot for matching-engine planning."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook


class WorkbookSnapshot:
    """Read only selected sheets and write sample rows to text.

    This is intentionally read-only. It helps us inspect real row patterns in
    Taxepunkter, Taxa_från_edp and Taxa_Saknas before we build matching logic.
    """

    DEFAULT_SHEETS = ["Taxepunkter", "Taxa_från_edp", "Taxa_Saknas", "Kontrollrapport"]

    def write_snapshot(
        self,
        workbook_path: str | Path,
        out_path: str | Path = "output/excel/arbets_excel_snapshot.txt",
        max_rows: int = 25,
    ) -> Path:
        source = Path(workbook_path)
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        lines: list[str] = [
            "Arbets-Excel Snapshot",
            "",
            f"Workbook: {source}",
            "This is read-only. Arbets-Excel is not master.",
            "",
        ]

        if not source.exists():
            lines.append(f"Missing workbook: {source}")
            out.write_text("\n".join(lines), encoding="utf-8")
            return out

        wb = load_workbook(source, data_only=False)

        for sheet_name in self.DEFAULT_SHEETS:
            if sheet_name not in wb.sheetnames:
                lines.append("=" * 100)
                lines.append(f"Sheet missing: {sheet_name}")
                lines.append("")
                continue

            ws = wb[sheet_name]
            lines.append("=" * 100)
            lines.append(f"Sheet: {sheet_name}")
            lines.append(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
            lines.append("")

            for row_idx in range(1, min(ws.max_row, max_rows) + 1):
                values = []
                for col_idx in range(1, min(ws.max_column, 12) + 1):
                    value = ws.cell(row_idx, col_idx).value
                    values.append("" if value is None else str(value))
                lines.append(f"{row_idx:04d}: " + " | ".join(values))
            lines.append("")

        out.write_text("\n".join(lines), encoding="utf-8")
        return out
