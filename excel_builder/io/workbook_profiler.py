"""Profile the current Arbets-Excel without modifying it.

This profiler is read-only. It is used before Excel Builder starts writing any
new workbook so we understand sheets, tables, hidden columns, formulas and data
validations.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_builder.models import (
    ColumnProfile,
    DataValidationProfile,
    SheetProfile,
    TableProfile,
    WorkbookProfile,
)


class WorkbookProfiler:
    FIELD_ALIASES = {
        "section": ["paragraf", "section", "avsnitt"],
        "paragraph_name": ["paragrafnamn", "rubrik", "paragraf namn"],
        "tax_point": ["taxapunkt", "taxepunkt", "taxa", "benämning", "benamning"],
        "variant": ["variant", "avgiftstyp", "intervall"],
        "unit": ["enhet", "unit"],
        "price": ["föreslagen taxa", "pris", "belopp", "avgift"],
        "edp_code": ["taxakod", "taxekod", "edp", "edp-kod"],
    }

    def profile(self, path: str | Path) -> WorkbookProfile:
        workbook_path = Path(path)
        profile = WorkbookProfile(path=str(workbook_path))

        if not workbook_path.exists():
            profile.warnings.append(f"Arbets-Excel saknas: {workbook_path}")
            return profile

        wb_values = load_workbook(workbook_path, data_only=True)
        wb_formulas = load_workbook(workbook_path, data_only=False)

        for ws_values in wb_values.worksheets:
            ws_formulas = wb_formulas[ws_values.title]
            sheet = SheetProfile(
                name=ws_values.title,
                max_row=ws_values.max_row,
                max_column=ws_values.max_column,
                hidden_state=ws_values.sheet_state,
                frozen_panes=str(ws_values.freeze_panes or ""),
                merged_ranges=[str(rng) for rng in ws_values.merged_cells.ranges],
            )

            sheet.likely_header_row = self._find_header_row(ws_values)
            sheet.columns = self._profile_columns(ws_values, ws_formulas, sheet.likely_header_row)
            sheet.detected_fields = self._detect_fields(sheet.columns)
            sheet.tables = self._profile_tables(ws_values)
            sheet.data_validations = self._profile_data_validations(ws_values)

            profile.sheets.append(sheet)

        return profile

    def _find_header_row(self, ws) -> int | None:
        best_row = None
        best_score = 0

        for row_idx in range(1, min(ws.max_row, 50) + 1):
            values = [
                self._norm(ws.cell(row_idx, col_idx).value)
                for col_idx in range(1, ws.max_column + 1)
            ]
            if not any(values):
                continue

            score = 0
            for value in values:
                for aliases in self.FIELD_ALIASES.values():
                    if value in aliases or any(alias in value for alias in aliases):
                        score += 1

            if score > best_score:
                best_score = score
                best_row = row_idx

        return best_row

    def _profile_columns(self, ws_values, ws_formulas, header_row: int | None) -> list[ColumnProfile]:
        columns: list[ColumnProfile] = []

        for col_idx in range(1, ws_values.max_column + 1):
            letter = get_column_letter(col_idx)
            header = ""
            if header_row:
                header = str(ws_values.cell(header_row, col_idx).value or "").strip()

            non_empty_count = 0
            formula_count = 0

            for row_idx in range(1, ws_values.max_row + 1):
                value = ws_values.cell(row_idx, col_idx).value
                formula_value = ws_formulas.cell(row_idx, col_idx).value

                if value not in (None, ""):
                    non_empty_count += 1
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    formula_count += 1

            dim = ws_values.column_dimensions.get(letter)

            columns.append(
                ColumnProfile(
                    index=col_idx,
                    letter=letter,
                    header=header,
                    hidden=bool(dim.hidden) if dim else False,
                    width=float(dim.width) if dim and dim.width else None,
                    non_empty_count=non_empty_count,
                    formula_count=formula_count,
                )
            )

        return columns

    def _profile_tables(self, ws) -> list[TableProfile]:
        tables: list[TableProfile] = []
        for table in ws.tables.values():
            columns = []
            try:
                columns = [col.name for col in table.tableColumns]
            except Exception:
                columns = []
            tables.append(TableProfile(name=table.name, ref=table.ref, columns=columns))
        return tables

    def _profile_data_validations(self, ws) -> list[DataValidationProfile]:
        result: list[DataValidationProfile] = []
        validations = getattr(ws.data_validations, "dataValidation", [])
        for dv in validations:
            result.append(
                DataValidationProfile(
                    type=str(dv.type or ""),
                    sqref=str(dv.sqref or ""),
                    formula1=str(dv.formula1 or ""),
                    formula2=str(dv.formula2 or ""),
                )
            )
        return result

    def _detect_fields(self, columns: list[ColumnProfile]) -> dict[str, str]:
        detected: dict[str, str] = {}

        for col in columns:
            header = self._norm(col.header)
            if not header:
                continue

            for field, aliases in self.FIELD_ALIASES.items():
                if field in detected:
                    continue
                if header in aliases or any(alias in header for alias in aliases):
                    detected[field] = col.letter

        return detected

    def _norm(self, value) -> str:
        return " ".join(str(value or "").replace("\xa0", " ").lower().strip().split())
