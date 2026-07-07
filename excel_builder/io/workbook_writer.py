"""Write Arbets-Excel output from parser rows using the immutable master template."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_builder.edp.proposal_trace_sheets import ProposalTraceSheets
from excel_builder.guards import ImmutableMasterGuard
from excel_builder.models import BuilderResult
from excel_builder.template import TemplateMasterManager


class WorkbookWriter:
    """Create a working copy from the selected master and write generated data.

    Protected source areas are deliberately not touched:
    - Taxepunkter columns A:E are template columns and immutable in the master.
    - Taxa_från_edp is EDP facit and immutable.

    Generated parser rows are written to a separate Builder_Output sheet until a
    later approved block maps them into allowed working columns.
    """

    OUTPUT_SHEET = "Builder_Output"
    HEADERS = [
        "Paragraf",
        "Paragrafnamn",
        "Taxapunkt",
        "Variant",
        "Enhet",
        "Pris från Word",
        "EDP taxekod",
        "EDP koppling status",
        "Kommentar",
        "Källa",
    ]

    def write(self, result: BuilderResult, path: str | Path = "output/excel/ArbetsExcel_byggd_fran_parser.xlsx") -> Path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)

        template_info = TemplateMasterManager().create_working_copy(out)
        if template_info.warnings:
            raise RuntimeError("; ".join(template_info.warnings))

        guard = ImmutableMasterGuard([template_info.template_path])
        template_fingerprint = guard.fingerprint(template_info.template_path)

        wb = load_workbook(out)
        ws = self._replace_sheet(wb, self.OUTPUT_SHEET)
        ws.append(self.HEADERS)

        for row in result.rows:
            ws.append([
                row.section,
                "",
                row.name,
                row.variant,
                row.unit,
                row.price,
                "",
                "Ej kopplad",
                "",
                row.source,
            ])

        self._style(ws)
        self._add_table(ws)
        self._replace_summary_sheet(wb, result, template_info.template_path)
        self._replace_readme_sheet(wb)
        ProposalTraceSheets().add(wb, municipality="", context="Parseroutput")

        wb.save(out)
        guard.verify_unchanged(template_fingerprint)
        return out

    def _replace_sheet(self, wb, sheet_name: str):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        return wb.create_sheet(sheet_name)

    def _style(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)

        widths = {
            "A": 14,
            "B": 28,
            "C": 55,
            "D": 32,
            "E": 16,
            "F": 18,
            "G": 18,
            "H": 22,
            "I": 35,
            "J": 24,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _add_table(self, ws) -> None:
        if ws.max_row < 2:
            return
        ref = f"A1:J{ws.max_row}"
        table = Table(displayName="BuilderOutputTable", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    def _replace_summary_sheet(self, wb, result: BuilderResult, template_path: str) -> None:
        ws = self._replace_sheet(wb, "Sammanfattning")
        ws.append(["Mått", "Värde"])
        ws.append(["Antal rader från parser", result.row_count])
        ws.append(["Varningar", len(result.warnings)])
        ws.append(["Status", "Arbets-Excel skapad som kopia av immutable master"])
        ws.append(["Master", template_path])
        ws.append(["Skydd", "Taxepunkter A:E och Taxa_från_edp skrivs inte automatiskt"])
        ws.append(["Genererad data", self.OUTPUT_SHEET])
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 80
        for cell in ws[1]:
            cell.font = Font(bold=True)

    def _replace_readme_sheet(self, wb) -> None:
        ws = self._replace_sheet(wb, "README")
        rows = [
            ["Excel Builder"],
            [""],
            ["Denna fil är en arbetskopia skapad från immutable Excel-master."],
            ["Masterfilen får aldrig ändras eller skrivas över."],
            ["Taxepunkter kolumn A:E är mallkolumner och lämnas orörda."],
            ["Taxa_från_edp är facit och lämnas helt orörd."],
            [""],
            ["Genererade parserrader skrivs till Builder_Output."],
            ["Om master behöver ändras ska en ny versionsfil skapas och godkännas."],
        ]
        for row in rows:
            ws.append(row)
        ws.column_dimensions["A"].width = 100
