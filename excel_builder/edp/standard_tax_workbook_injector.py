"""Attach EDP standard tax reference sheets to generated workbooks.

The standard tax workbook is reference knowledge. It may be used in every
project, but it must never mix municipality-specific EDP data between projects.
"""

from __future__ import annotations

from pathlib import Path
from copy import copy

from openpyxl import load_workbook


class StandardTaxWorkbookInjector:
    DEFAULT_PATH = Path("data/edp_standard/EDP_Future_Standard_Taxor_Renhallning.xlsx")
    SHEET_PREFIX = "EDP_Standard_"

    def attach(self, target_wb, source_path: str | Path | None = None) -> list[str]:
        source = Path(source_path) if source_path else self.DEFAULT_PATH
        if not source.exists():
            return [f"Standardtaxefil saknas: {source}"]

        source_wb = load_workbook(source, data_only=False, read_only=False)
        warnings: list[str] = []

        for source_ws in source_wb.worksheets:
            target_title = self._safe_title(f"{self.SHEET_PREFIX}{source_ws.title}", target_wb.sheetnames)
            target_ws = target_wb.create_sheet(target_title)

            self._copy_sheet_content(source_ws, target_ws)

        return warnings

    def _safe_title(self, title: str, existing: list[str]) -> str:
        # Excel sheet names max 31 chars.
        base = title[:31]
        if base not in existing:
            return base

        counter = 2
        while True:
            suffix = f"_{counter}"
            candidate = f"{base[:31-len(suffix)]}{suffix}"
            if candidate not in existing:
                return candidate
            counter += 1

    def _copy_sheet_content(self, source_ws, target_ws) -> None:
        for row in source_ws.iter_rows():
            for source_cell in row:
                target_cell = target_ws.cell(source_cell.row, source_cell.column, source_cell.value)

                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)

                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)

        for col_letter, dim in source_ws.column_dimensions.items():
            target_dim = target_ws.column_dimensions[col_letter]
            target_dim.width = dim.width
            target_dim.hidden = dim.hidden

        for row_idx, dim in source_ws.row_dimensions.items():
            target_dim = target_ws.row_dimensions[row_idx]
            target_dim.height = dim.height
            target_dim.hidden = dim.hidden

        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

        target_ws.freeze_panes = source_ws.freeze_panes
        target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
