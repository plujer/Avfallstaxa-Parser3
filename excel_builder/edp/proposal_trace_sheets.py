"""Add proposal, deviation and rule tracing sheets to generated workbooks."""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


class ProposalTraceSheets:
    """Create standard review sheets in every generated workbook.

    Taxa_Förslag:
        Holds suggested tax codes when no safe municipality-specific EDP match exists.

    EDP_Export_Avviker_Standard:
        Holds detected deviations where existing EDP export differs from standard tax
        reference. Existing Taxa_från_edp codes/prices remain fixed and must not be
        overwritten automatically.

    Regelspårning:
        Documents how each generated row/value was created.
    """

    PROPOSAL_SHEET = "Taxa_Förslag"
    DEVIATION_SHEET = "EDP_Avviker_Standard"
    TRACE_SHEET = "Regelspårning"

    PROPOSAL_HEADERS = [
        "Kommun",
        "Paragraf",
        "Taxapunkt",
        "Variant",
        "Enhet",
        "Föreslagen taxekod",
        "Föreslagen benämning",
        "Källa",
        "Säkerhet",
        "Status",
        "Kommentar",
    ]

    DEVIATION_HEADERS = [
        "Kommun",
        "EDP taxekod",
        "EDP benämning",
        "EDP faktor",
        "EDP taxedel",
        "Standard taxekod",
        "Standard benämning",
        "Standard faktor",
        "Standard taxedel",
        "Avvikelse",
        "Rekommendation",
        "Status",
    ]

    TRACE_HEADERS = [
        "Kommun",
        "Paragraf",
        "Taxapunkt",
        "Fält",
        "Värde",
        "Ursprung",
        "Regel",
        "Källa",
        "Status",
        "Kommentar",
    ]

    def add(self, wb, municipality: str = "", context: str = "") -> None:
        self._remove_existing_if_empty_or_template(wb, self.PROPOSAL_SHEET)
        self._remove_existing_if_empty_or_template(wb, self.DEVIATION_SHEET)
        self._remove_existing_if_empty_or_template(wb, self.TRACE_SHEET)

        proposal = wb.create_sheet(self.PROPOSAL_SHEET)
        deviation = wb.create_sheet(self.DEVIATION_SHEET)
        trace = wb.create_sheet(self.TRACE_SHEET)

        self._setup_proposal_sheet(proposal, municipality, context)
        self._setup_deviation_sheet(deviation, municipality, context)
        self._setup_trace_sheet(trace, municipality, context)

    def _setup_proposal_sheet(self, ws, municipality: str, context: str) -> None:
        ws.append(["Taxa_Förslag"])
        ws.append(["Syfte", "Förslag på saknade taxekoder. Förslag ska granskas innan import till EDP."])
        ws.append(["Kommun", municipality])
        ws.append(["Källa", context])
        ws.append([])
        ws.append(self.PROPOSAL_HEADERS)

        ws.append([
            municipality,
            "",
            "",
            "",
            "",
            "",
            "",
            "EDP Standardtaxor / regelverk / manuell granskning",
            "",
            "Ej granskad",
            "EDP-data får aldrig blandas mellan projekt.",
        ])

        self._style(ws, header_row=6, table_name="TaxaForslagTable")
        self._set_widths(ws, {
            "A": 16, "B": 12, "C": 48, "D": 28, "E": 14,
            "F": 22, "G": 44, "H": 42, "I": 12, "J": 18, "K": 50,
        })

    def _setup_deviation_sheet(self, ws, municipality: str, context: str) -> None:
        ws.append(["EDP Export avviker från standardtaxa"])
        ws.append(["Syfte", "Visar avvikelser mot standardtaxor. Befintlig Taxa_från_edp är fast och ändras aldrig automatiskt."])
        ws.append(["Kommun", municipality])
        ws.append(["Källa", context])
        ws.append([])
        ws.append(self.DEVIATION_HEADERS)

        ws.append([
            municipality,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Granska manuellt. Ändra inte befintlig EDP automatiskt.",
            "Ej granskad",
        ])

        self._style(ws, header_row=6, table_name="EdpAvvikerStandardTable")
        self._set_widths(ws, {
            "A": 16, "B": 18, "C": 44, "D": 18, "E": 42,
            "F": 22, "G": 44, "H": 18, "I": 42, "J": 46,
            "K": 54, "L": 18,
        })

    def _setup_trace_sheet(self, ws, municipality: str, context: str) -> None:
        ws.append(["Regelspårning"])
        ws.append(["Syfte", "Spårar hur varje rad/fält byggdes: Word, EDP, standardtaxor eller manuell granskning."])
        ws.append(["Kommun", municipality])
        ws.append(["Källa", context])
        ws.append([])
        ws.append(self.TRACE_HEADERS)

        ws.append([
            municipality,
            "",
            "",
            "Taxakod",
            "",
            "Ej satt",
            "Befintlig Taxa_från_edp är fast. Standardtaxor får endast ge förslag/avvikelse.",
            "Regelverk",
            "Ej granskad",
            "Standardflik för framtida spårning.",
        ])

        self._style(ws, header_row=6, table_name="RegelsparningTable")
        self._set_widths(ws, {
            "A": 16, "B": 12, "C": 48, "D": 18, "E": 28,
            "F": 24, "G": 70, "H": 30, "I": 18, "J": 50,
        })

    def _style(self, ws, header_row: int, table_name: str) -> None:
        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        note_fill = PatternFill("solid", fgColor="FFF2CC")

        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = title_fill

        for row_idx in [2, 3, 4]:
            ws.cell(row_idx, 1).font = Font(bold=True)
            ws.cell(row_idx, 1).fill = note_fill
            ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True)

        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = f"A{header_row + 1}"

        max_col_letter = ws.cell(header_row, ws.max_column).column_letter
        ref = f"A{header_row}:{max_col_letter}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    def _set_widths(self, ws, widths: dict[str, int]) -> None:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

    def _remove_existing_if_empty_or_template(self, wb, sheet_name: str) -> None:
        if sheet_name not in wb.sheetnames:
            return
        wb.remove(wb[sheet_name])
