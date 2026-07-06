"""Read one EDP export file in isolation.

Each EDP export belongs to one municipality/run and must never be mixed with
another municipality's workbook or Word output.
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from excel_builder.models import EdpExport, EdpExportRow


class EdpExportReader:
    REQUIRED_HEADERS = [
        "intRecnum",
        "strTaxekod",
        "strTaxebenamning",
        "strProdukt",
        "strDelProdukt",
        "bytDelradnr",
        "strFaktor",
        "strTaxedelAvser",
        "strEntreprenorkod",
        "strRenhDistrKod",
        "curNuvarandePris",
        "datNuvarandePrisDatum",
        "bolPrisPerTomning",
        "strAvvikandeFormel",
        "bolPriserInklMoms",
        "bytMomskod",
        "strFormel",
    ]

    def read(self, path: str | Path, municipality: str) -> EdpExport:
        source = Path(path)
        export = EdpExport(municipality=municipality, source_path=str(source))

        if not source.exists():
            export.warnings.append(f"EDP-export saknas: {source}")
            return export

        wb = load_workbook(source, data_only=True, read_only=False)
        ws = wb[wb.sheetnames[0]]

        header_row = self._find_header_row(ws)
        if header_row is None:
            export.warnings.append("Kunde inte hitta EDP-header med intRecnum/strTaxekod.")
            return export

        headers = [str(ws.cell(header_row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        header_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

        missing = [header for header in self.REQUIRED_HEADERS if header not in header_index]
        if missing:
            export.warnings.append(f"Saknade EDP-kolumner: {', '.join(missing)}")

        for row_idx in range(header_row + 1, ws.max_row + 1):
            values = {header: self._value(ws.cell(row_idx, header_index[header]).value) for header in self.REQUIRED_HEADERS if header in header_index}
            if not any(values.values()):
                continue
            export.rows.append(EdpExportRow(**values))

        return export

    def _find_header_row(self, ws) -> int | None:
        for row_idx in range(1, min(ws.max_row, 20) + 1):
            values = [str(ws.cell(row_idx, col).value or "").strip() for col in range(1, min(ws.max_column, 25) + 1)]
            if "intRecnum" in values and "strTaxekod" in values:
                return row_idx
        return None

    def _value(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()
