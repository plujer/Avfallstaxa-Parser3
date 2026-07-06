"""Build one isolated output workbook for one EDP export.

This does not create the final production workbook yet. It creates a controlled
run artifact proving that municipality/EDP data is isolated and written to its
own workbook.
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_builder.edp.standard_tax_workbook_injector import StandardTaxWorkbookInjector
from excel_builder.models import EdpExport


class IsolatedWorkbookBuilder:
    EDP_HEADERS = [
        "intRecnum",
        "strTaxekod",
        "strTaxebenamning",
        "strProdukt",
        "strDelProdukt",
        "bytDelradnr",
        "strFaktor",
        "strTaxedelAvser",
        "strEntreprenorkod",
        "strRenhDistrKod",
        "curNuvarandePris",
        "datNuvarandePrisDatum",
        "bolPrisPerTomning",
        "strAvvikandeFormel",
        "bolPriserInklMoms",
        "bytMomskod",
        "strFormel",
    ]

    def __init__(self, include_standard_tax_sheets: bool = True) -> None:
        self.include_standard_tax_sheets = include_standard_tax_sheets
        self.standard_injector = StandardTaxWorkbookInjector()

    def build(self, edp_export: EdpExport, out_path: str | Path) -> Path:
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Taxa_från_edp"

        ws.append(self.EDP_HEADERS)
        for row in edp_export.rows:
            ws.append([getattr(row, header) for header in self.EDP_HEADERS])

        self._style_edp_sheet(ws)

        info = wb.create_sheet("Körningsinfo")
        info.append(["Fält", "Värde"])
        info.append(["Kommun", edp_export.municipality])
        info.append(["EDP-källa", edp_export.source_path])
        info.append(["EDP-rader", edp_export.row_count])
        info.append(["Status", "Isolerad körning – får inte blandas med annan kommun"])
        info.append(["Standardtaxor", "Inkluderade som referensflikar om standardtaxefilen finns"])
        info.column_dimensions["A"].width = 24
        info.column_dimensions["B"].width = 90
        for cell in info[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        if self.include_standard_tax_sheets:
            warnings = self.standard_injector.attach(wb)
            if warnings:
                start_row = info.max_row + 2
                info.cell(start_row, 1).value = "Standardtaxevarningar"
                info.cell(start_row, 1).font = Font(bold=True)
                for idx, warning in enumerate(warnings, start=start_row + 1):
                    info.cell(idx, 1).value = warning

        wb.save(out)
        return out

    def _style_edp_sheet(self, ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        ws.freeze_panes = "A2"
        widths = {
            "A": 14, "B": 18, "C": 42, "D": 14, "E": 16, "F": 12, "G": 14,
            "H": 52, "I": 16, "J": 16, "K": 18, "L": 18, "M": 16, "N": 20,
            "O": 16, "P": 12, "Q": 24,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        if ws.max_row > 1:
            table = Table(displayName="TaxaFranEdpTable", ref=f"A1:Q{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
