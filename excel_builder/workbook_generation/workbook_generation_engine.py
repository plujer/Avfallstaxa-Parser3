"""Write decision support and traceability into generated Arbets-Excel workbooks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_builder.models import WorkbookDecisionRow, WorkbookGenerationReport


class WorkbookGenerationEngine:
    """Create a workbook-level decision trace without changing Taxa_från_edp.

    The engine writes review/support sheets only. Existing EDP data remains fixed.
    """

    TRACE_SHEET = "Decision_Trace"
    REPORT_SHEET = "Workbook_Generation"
    TRACE_HEADERS = [
        "Word taxekod",
        "Kandidat taxekod",
        "Beslut",
        "Säkerhet",
        "Totalpoäng",
        "Primär motivering",
        "Avvisningsorsak",
        "Signaler",
    ]
    TAXEPUNKTER_HEADERS = [
        "Beslutsspår kandidat",
        "Beslutsspår status",
        "Beslutsspår säkerhet",
        "Beslutsspår motivering",
    ]

    def write(self, workbook_path: str | Path, decisions: list[WorkbookDecisionRow]) -> WorkbookGenerationReport:
        target = Path(workbook_path)
        report = WorkbookGenerationReport(workbook_path=str(target))
        if not target.exists():
            report.warnings.append(f"Workbook saknas: {target}")
            return report

        wb = load_workbook(target)
        self._replace_sheet(wb, self.TRACE_SHEET)
        self._replace_sheet(wb, self.REPORT_SHEET)
        trace_ws = wb.create_sheet(self.TRACE_SHEET)
        report_ws = wb.create_sheet(self.REPORT_SHEET)

        report.rows_written = self._write_decision_trace(trace_ws, decisions)
        if "Taxepunkter" in wb.sheetnames:
            report.taxepunkter_rows_updated = self._write_taxepunkter_trace(wb["Taxepunkter"], decisions)
        else:
            report.warnings.append("Taxepunkter saknas i arbetsboken.")

        self._write_report(report_ws, report, len(decisions))
        wb.save(target)
        return report

    def _write_decision_trace(self, ws, decisions: list[WorkbookDecisionRow]) -> int:
        ws.append(["Decision_Trace"])
        ws.append(["Syfte", "Spårbar beslutskedja från Word/parser till kandidat och beslutsmotivering."])
        ws.append(["Regel", "Taxa_från_edp är facit och ändras aldrig automatiskt."])
        ws.append([])
        ws.append(self.TRACE_HEADERS)
        for row in decisions:
            ws.append([
                row.word_tax_code,
                row.candidate_tax_code,
                row.decision,
                row.confidence,
                row.total_score,
                row.primary_reason,
                row.rejected_reason,
                row.signals,
            ])
        self._style_table(ws, header_row=5, table_name="DecisionTraceTable")
        self._set_widths(ws, {"A": 20, "B": 22, "C": 14, "D": 12, "E": 12, "F": 60, "G": 45, "H": 90})
        return len(decisions)

    def _write_taxepunkter_trace(self, ws, decisions: list[WorkbookDecisionRow]) -> int:
        if ws.max_row < 2 or not decisions:
            return 0
        header_map = {str(ws.cell(1, col).value or "").strip(): col for col in range(1, ws.max_column + 1)}
        start_col = ws.max_column + 1
        for idx, header in enumerate(self.TAXEPUNKTER_HEADERS):
            col = header_map.get(header, start_col + idx)
            ws.cell(1, col).value = header
            ws.cell(1, col).font = Font(bold=True)
            ws.cell(1, col).fill = PatternFill("solid", fgColor="D9EAF7")
            header_map[header] = col

        updated = 0
        for row_idx, decision in enumerate(decisions, start=2):
            if row_idx > ws.max_row:
                break
            ws.cell(row_idx, header_map["Beslutsspår kandidat"]).value = decision.candidate_tax_code
            ws.cell(row_idx, header_map["Beslutsspår status"]).value = decision.decision
            ws.cell(row_idx, header_map["Beslutsspår säkerhet"]).value = decision.confidence
            ws.cell(row_idx, header_map["Beslutsspår motivering"]).value = decision.primary_reason
            self._apply_decision_fill(ws, row_idx, header_map["Beslutsspår status"], decision.decision)
            updated += 1

        for header, width in {
            "Beslutsspår kandidat": 24,
            "Beslutsspår status": 18,
            "Beslutsspår säkerhet": 18,
            "Beslutsspår motivering": 60,
        }.items():
            ws.column_dimensions[ws.cell(1, header_map[header]).column_letter].width = width
        return updated

    def _write_report(self, ws, report: WorkbookGenerationReport, input_count: int) -> None:
        rows = [
            ["Workbook Generation Engine Report"],
            [],
            ["Status", report.status],
            ["Workbook", report.workbook_path],
            ["Decision trace input rows", input_count],
            ["Decision_Trace rows written", report.rows_written],
            ["Taxepunkter rows updated", report.taxepunkter_rows_updated],
            ["Warnings", len(report.warnings)],
            [],
            ["Regel", "Taxa_från_edp ändras inte. Beslutsspår är beslutsstöd för granskning."],
        ]
        if report.warnings:
            rows.append([])
            rows.append(["Warnings"])
            for warning in report.warnings:
                rows.append([warning])
        for row in rows:
            ws.append(row)
        ws["A1"].font = Font(bold=True, size=14)
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 90

    def _replace_sheet(self, wb, sheet_name: str) -> None:
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

    def _style_table(self, ws, header_row: int, table_name: str) -> None:
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        for row_idx in [2, 3]:
            ws.cell(row_idx, 1).font = Font(bold=True)
            ws.cell(row_idx, 1).fill = PatternFill("solid", fgColor="FFF2CC")
            ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True)
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(wrap_text=True)
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = f"A{header_row + 1}"
        if ws.max_row >= header_row:
            ref = f"A{header_row}:{ws.cell(header_row, ws.max_column).column_letter}{ws.max_row}"
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
            ws.add_table(table)

    def _set_widths(self, ws, widths: dict[str, int]) -> None:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

    def _apply_decision_fill(self, ws, row_idx: int, col_idx: int, decision: str) -> None:
        colors = {"ACCEPT": "C6EFCE", "REVIEW": "FFF2CC", "REJECT": "F4CCCC"}
        ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=colors.get(decision, "E7E6E6"))
