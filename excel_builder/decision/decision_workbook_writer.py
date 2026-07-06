"""Write consolidated decisions to workbook sheets."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from excel_builder.edp import ProposalTraceSheets
from excel_builder.models import TaxDecisionReport


class DecisionWorkbookWriter:
    def write(self, workbook_path: str | Path, report: TaxDecisionReport, municipality: str = "") -> Path:
        target = Path(workbook_path)
        wb = load_workbook(target)

        if "Taxepunkter" in wb.sheetnames:
            self._write_taxepunkter_status(wb["Taxepunkter"], report)

        if "Taxa_Förslag" not in wb.sheetnames or "Regelspårning" not in wb.sheetnames:
            ProposalTraceSheets().add(wb, municipality=municipality, context="Beslutsmotor")

        self._write_trace(wb["Regelspårning"], report, municipality)
        self._write_proposals(wb["Taxa_Förslag"], report, municipality)

        wb.save(target)
        return target

    def _write_taxepunkter_status(self, ws, report: TaxDecisionReport) -> None:
        header_row = 1
        status_col = ws.max_column + 1
        rule_col = ws.max_column + 2
        comment_col = ws.max_column + 3

        existing_headers = {str(ws.cell(header_row, col).value or ""): col for col in range(1, ws.max_column + 1)}
        status_col = existing_headers.get("Beslutsstatus", status_col)
        rule_col = existing_headers.get("Beslutsregel", rule_col)
        comment_col = existing_headers.get("Beslutskommentar", comment_col)

        ws.cell(header_row, status_col).value = "Beslutsstatus"
        ws.cell(header_row, rule_col).value = "Beslutsregel"
        ws.cell(header_row, comment_col).value = "Beslutskommentar"

        for col in [status_col, rule_col, comment_col]:
            ws.cell(header_row, col).font = Font(bold=True)
            ws.cell(header_row, col).fill = PatternFill("solid", fgColor="D9EAF7")

        # Current generated Taxepunkter has parser rows in the same order starting row 2.
        for idx, decision in enumerate(report.decisions, start=2):
            if idx > ws.max_row:
                break
            ws.cell(idx, status_col).value = decision.status
            ws.cell(idx, rule_col).value = decision.rule
            ws.cell(idx, comment_col).value = decision.comment

        ws.column_dimensions[ws.cell(header_row, status_col).column_letter].width = 22
        ws.column_dimensions[ws.cell(header_row, rule_col).column_letter].width = 48
        ws.column_dimensions[ws.cell(header_row, comment_col).column_letter].width = 60

    def _write_trace(self, ws, report: TaxDecisionReport, municipality: str) -> None:
        self._remove_template_rows(ws)

        for decision in report.decisions:
            ws.append([
                municipality,
                decision.parser_row.section,
                decision.parser_row.tax_point,
                "Beslut",
                decision.status,
                decision.source,
                decision.rule,
                "TaxDecisionEngine",
                decision.status,
                decision.comment,
            ])

    def _write_proposals(self, ws, report: TaxDecisionReport, municipality: str) -> None:
        self._remove_template_rows(ws)

        for decision in report.decisions:
            if decision.status != "STANDARD_PROPOSAL" or not decision.standard_row:
                continue

            standard = decision.standard_row
            ws.append([
                municipality,
                decision.parser_row.section,
                decision.parser_row.tax_point,
                decision.parser_row.variant,
                decision.parser_row.unit,
                standard.strTaxekod,
                standard.strTaxebenamning,
                f"Standardtaxor: {standard.source_sheet} rad {standard.row_number}",
                f"{decision.confidence:.3f}",
                "Ej granskad",
                decision.comment,
            ])

    def _remove_template_rows(self, ws) -> None:
        if ws.max_row >= 7:
            row_values = [str(ws.cell(7, col).value or "") for col in range(1, ws.max_column + 1)]
            joined = " ".join(row_values).lower()
            if "ej granskad" in joined and (
                "standardflik" in joined
                or "edp-data får aldrig blandas" in joined
                or "granska manuellt" in joined
            ):
                ws.delete_rows(7, 1)
