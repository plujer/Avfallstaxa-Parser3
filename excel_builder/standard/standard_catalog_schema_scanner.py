"""Reverse engineer the developer standard tax workbook.

The standard tax file may contain several logical sections per sheet. This
scanner tries to detect every section with tax-code-like headers, rather than
assuming one table per sheet.
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from excel_builder.matching import MatchNormalizer
from excel_builder.models import StandardCatalogSchema, StandardCatalogSection, StandardCatalogSheetSchema


class StandardCatalogSchemaScanner:
    MAX_SCAN_COLS = 80

    HEADER_TERMS = {
        "strtaxekod",
        "taxekod",
        "taxakod",
        "strtaxebenamning",
        "taxebenämning",
        "taxebenamning",
        "benämning",
        "benamning",
        "strfaktor",
        "faktor",
        "strtaxedelavser",
        "taxedel",
        "strformel",
        "formel",
    }

    def __init__(self) -> None:
        self.normalizer = MatchNormalizer()

    def scan(self, path: str | Path) -> StandardCatalogSchema:
        source = Path(path)
        schema = StandardCatalogSchema(source_path=str(source))

        if not source.exists():
            schema.warnings.append(f"Standardtaxefil saknas: {source}")
            return schema

        wb = load_workbook(source, data_only=True, read_only=False)

        for ws in wb.worksheets:
            sheet = StandardCatalogSheetSchema(
                name=ws.title,
                max_row=ws.max_row,
                max_column=ws.max_column,
                tables=sorted(list(ws.tables.keys())),
                merged_ranges=[str(item) for item in ws.merged_cells.ranges],
                formula_count=self._formula_count(ws),
            )
            sheet.sections = self._detect_sections(ws)
            schema.sheets.append(sheet)

        return schema

    def _detect_sections(self, ws) -> list[StandardCatalogSection]:
        header_rows: list[int] = []

        for row_idx in range(1, ws.max_row + 1):
            values = [self._value(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, self.MAX_SCAN_COLS) + 1)]
            score = self._header_score(values)
            if score >= 0.45:
                header_rows.append(row_idx)

        sections: list[StandardCatalogSection] = []

        for idx, header_row in enumerate(header_rows):
            next_header = header_rows[idx + 1] if idx + 1 < len(header_rows) else ws.max_row + 1
            start_row = header_row + 1
            end_row = self._find_section_end(ws, start_row, next_header - 1)
            headers = [self._value(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
            key_columns = self._key_columns(headers)
            row_count = max(0, end_row - start_row + 1)

            if row_count <= 0:
                continue

            sections.append(
                StandardCatalogSection(
                    sheet_name=ws.title,
                    header_row=header_row,
                    start_row=start_row,
                    end_row=end_row,
                    headers=headers,
                    row_count=row_count,
                    key_columns=key_columns,
                )
            )

        return sections

    def _find_section_end(self, ws, start_row: int, max_row: int) -> int:
        last_data_row = start_row - 1
        blank_streak = 0

        for row_idx in range(start_row, max_row + 1):
            values = [self._value(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, self.MAX_SCAN_COLS) + 1)]
            non_empty = [value for value in values if value]

            if non_empty:
                last_data_row = row_idx
                blank_streak = 0
            else:
                blank_streak += 1

            if blank_streak >= 3 and last_data_row >= start_row:
                break

        return last_data_row

    def _header_score(self, values: list[str]) -> float:
        non_empty = [value for value in values if value]
        if len(non_empty) < 2:
            return 0.0

        normalized = [self.normalizer.normalize(value) for value in non_empty]
        joined = " | ".join(normalized)

        hits = sum(1 for term in self.HEADER_TERMS if term in joined)
        score = 0.0
        score += min(hits / 4, 1.0) * 0.75
        score += min(len(non_empty) / 12, 1.0) * 0.25
        return round(score, 4)

    def _key_columns(self, headers: list[str]) -> list[str]:
        result = []
        for header in headers:
            norm = self.normalizer.normalize(header)
            if not norm:
                continue
            if any(term in norm for term in self.HEADER_TERMS):
                result.append(header)
        return result

    def _formula_count(self, ws) -> int:
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    count += 1
        return count

    def _value(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()
