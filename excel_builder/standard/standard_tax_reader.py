"""Read EDP standard tax catalog from developer reference workbook.

This reader now uses StandardCatalogSchemaScanner and can read multiple logical
sections per sheet. The standard catalog is global reference knowledge only.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from excel_builder.models import StandardTaxCatalog, StandardTaxRow
from excel_builder.standard.standard_catalog_schema_scanner import StandardCatalogSchemaScanner


class StandardTaxReader:
    DEFAULT_PATH = Path("data/edp_standard/EDP_Future_Standard_Taxor_Renhallning.xlsx")

    FIELD_ALIASES = {
        "strTaxekod": ["strtaxekod", "taxekod", "taxakod"],
        "strTaxebenamning": ["strtaxebenamning", "benämning", "benamning", "taxebenämning", "taxebenamning"],
        "strFaktor": ["strfaktor", "faktor"],
        "strTaxedelAvser": ["strtaxedelavser", "taxedel avser", "taxedel"],
        "strFormel": ["strformel", "formel"],
        "curNuvarandePris": ["curnuvarandepris", "pris", "nuvarande pris"],
    }

    def __init__(self) -> None:
        self.schema_scanner = StandardCatalogSchemaScanner()

    def read(self, path: str | Path | None = None) -> StandardTaxCatalog:
        source = Path(path) if path else self.DEFAULT_PATH
        catalog = StandardTaxCatalog(source_path=str(source))

        if not source.exists():
            catalog.warnings.append(f"Standardtaxefil saknas: {source}")
            return catalog

        schema = self.schema_scanner.scan(source)
        catalog.warnings.extend(schema.warnings)

        wb = load_workbook(source, data_only=True, read_only=False)

        for sheet_schema in schema.sheets:
            ws = wb[sheet_schema.name]

            if not sheet_schema.sections:
                catalog.warnings.append(f"Inga standardtaxesektioner hittades i flik: {sheet_schema.name}")
                continue

            for section in sheet_schema.sections:
                header_map = self._header_map(ws, section.header_row)
                if "strTaxekod" not in header_map and "strTaxebenamning" not in header_map:
                    catalog.warnings.append(f"Sektion saknar taxekod/benämning: {sheet_schema.name} rad {section.header_row}")
                    continue

                for row_idx in range(section.start_row, section.end_row + 1):
                    raw: dict[str, str] = {}
                    for field, col_idx in header_map.items():
                        raw[field] = self._value(ws.cell(row_idx, col_idx).value)

                    if not any(raw.values()):
                        continue

                    # Require at least a code or a meaningful name.
                    if not raw.get("strTaxekod") and len(raw.get("strTaxebenamning", "")) < 3:
                        continue

                    catalog.rows.append(
                        StandardTaxRow(
                            source_sheet=sheet_schema.name,
                            row_number=row_idx,
                            strTaxekod=raw.get("strTaxekod", ""),
                            strTaxebenamning=raw.get("strTaxebenamning", ""),
                            strFaktor=raw.get("strFaktor", ""),
                            strTaxedelAvser=raw.get("strTaxedelAvser", ""),
                            strFormel=raw.get("strFormel", ""),
                            curNuvarandePris=raw.get("curNuvarandePris", ""),
                            raw=raw,
                        )
                    )

        return catalog

    def _header_map(self, ws, header_row: int) -> dict[str, int]:
        result: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            value = self._norm(ws.cell(header_row, col_idx).value)
            if not value:
                continue
            for field, aliases in self.FIELD_ALIASES.items():
                if field in result:
                    continue
                if value in aliases or any(alias in value for alias in aliases):
                    result[field] = col_idx
        return result

    def _norm(self, value) -> str:
        return " ".join(str(value or "").replace("\xa0", " ").lower().strip().split())

    def _value(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()
