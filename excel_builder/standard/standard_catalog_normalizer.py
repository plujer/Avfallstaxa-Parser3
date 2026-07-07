"""Create a normalized standard tax workbook.

The user has allowed formatting/normalizing the standard tax file if it helps.
This creates a new generated normalized file; it does not overwrite the original.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_builder.models import StandardTaxCatalog


class StandardCatalogNormalizer:
    HEADERS = [
        "source_sheet",
        "source_row",
        "strTaxekod",
        "strTaxebenamning",
        "strFaktor",
        "strTaxedelAvser",
        "strFormel",
        "curNuvarandePris",
    ]

    def write(self, catalog: StandardTaxCatalog, path: str | Path = "output/excel/EDP_Standardtaxor_normalized.xlsx") -> Path:
        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Standardtaxor_Normalized"

        ws.append(self.HEADERS)
        for row in catalog.rows:
            ws.append([
                row.source_sheet,
                row.row_number,
                row.strTaxekod,
                row.strTaxebenamning,
                row.strFaktor,
                row.strTaxedelAvser,
                row.strFormel,
                row.curNuvarandePris,
            ])

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        widths = {
            "A": 28, "B": 12, "C": 18, "D": 48, "E": 18, "F": 48, "G": 28, "H": 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"

        if ws.max_row > 1:
            table = Table(displayName="StandardtaxorNormalizedTable", ref=f"A1:H{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        info = wb.create_sheet("README")
        info.append(["Fält", "Värde"])
        info.append(["Källa", catalog.source_path])
        info.append(["Rader", catalog.row_count])
        info.append(["Status", "Genererad normaliserad kopia. Originalfilen ändras inte."])
        for cell in info[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        info.column_dimensions["A"].width = 24
        info.column_dimensions["B"].width = 90

        wb.save(out)
        return out
