"""Write EDP-vs-standard deviations to generated workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from excel_builder.edp import ProposalTraceSheets
from excel_builder.standard.edp_standard_deviation_engine import EdpStandardDeviationReport


class DeviationWorkbookWriter:
    def write(self, workbook_path: str | Path, report: EdpStandardDeviationReport) -> Path:
        target = Path(workbook_path)
        wb = load_workbook(target)

        if "EDP_Avviker_Standard" not in wb.sheetnames or "Regelspårning" not in wb.sheetnames:
            ProposalTraceSheets().add(wb, municipality=report.municipality, context="EDP-standardavvikelse")

        deviation_ws = wb["EDP_Avviker_Standard"]
        trace_ws = wb["Regelspårning"]

        self._remove_template_rows(deviation_ws)
        self._remove_template_rows(trace_ws)

        for item in report.deviations:
            if item.status != "REVIEW":
                continue

            standard = item.standard_row
            deviation_ws.append([
                report.municipality,
                item.edp_row.strTaxekod,
                item.edp_row.strTaxebenamning,
                item.edp_row.strFaktor,
                item.edp_row.strTaxedelAvser,
                standard.strTaxekod if standard else "",
                standard.strTaxebenamning if standard else "",
                standard.strFaktor if standard else "",
                standard.strTaxedelAvser if standard else "",
                item.deviation_type,
                item.recommendation,
                "Ej granskad",
            ])

            trace_ws.append([
                report.municipality,
                "",
                item.edp_row.strTaxebenamning,
                "EDP_Avviker_Standard",
                item.edp_row.strTaxekod,
                "Kommun-EDP + Standardtaxor",
                "Befintlig Taxa_från_edp är fast. Standardtaxa används endast för avvikelseanalys.",
                f"EDP {item.edp_row.strTaxekod}",
                "REVIEW",
                item.deviation_type,
            ])

        wb.save(target)
        return target

    def _remove_template_rows(self, ws) -> None:
        if ws.max_row >= 7:
            row_values = [str(ws.cell(7, col).value or "") for col in range(1, ws.max_column + 1)]
            joined = " ".join(row_values).lower()
            if "ej granskad" in joined and (
                "granska manuellt" in joined
                or "standardflik" in joined
                or "edp-data får aldrig blandas" in joined
            ):
                ws.delete_rows(7, 1)
