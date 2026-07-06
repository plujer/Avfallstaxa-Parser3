"""Write standard tax suggestions into generated workbooks.

Rules:
- Never overwrite Taxa_från_edp.
- Never treat a standard tax suggestion as confirmed EDP.
- Write suggestions only to Taxa_Förslag.
- Write decision trace only to Regelspårning.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from excel_builder.edp import ProposalTraceSheets
from excel_builder.models import StandardTaxSuggestionReport


class SuggestionWorkbookWriter:
    def write(
        self,
        workbook_path: str | Path,
        report: StandardTaxSuggestionReport,
        municipality: str = "",
    ) -> Path:
        target = Path(workbook_path)
        wb = load_workbook(target)

        if "Taxa_Förslag" not in wb.sheetnames or "Regelspårning" not in wb.sheetnames:
            ProposalTraceSheets().add(wb, municipality=municipality, context="Standardtaxeförslag")

        proposal_ws = wb["Taxa_Förslag"]
        trace_ws = wb["Regelspårning"]

        # Remove template rows only if they are still the original example rows.
        self._remove_template_rows(proposal_ws)
        self._remove_template_rows(trace_ws)

        for item in report.suggestions:
            if item.status not in {"PROPOSAL", "REVIEW"}:
                continue

            standard = item.standard_row
            if standard is None:
                continue

            proposal_ws.append([
                municipality,
                item.parser_row.section,
                item.parser_row.tax_point,
                item.parser_row.variant,
                item.parser_row.unit,
                standard.strTaxekod,
                standard.strTaxebenamning,
                f"Standardtaxor: {standard.source_sheet} rad {standard.row_number}",
                f"{item.score:.3f}",
                "Ej granskad",
                item.comment,
            ])

            trace_ws.append([
                municipality,
                item.parser_row.section,
                item.parser_row.tax_point,
                "Taxakod",
                standard.strTaxekod,
                "Standardtaxa-förslag",
                "Standardtaxa får endast föreslås. Befintlig Taxa_från_edp ändras inte automatiskt.",
                f"{standard.source_sheet} rad {standard.row_number}",
                item.status,
                item.comment,
            ])

        wb.save(target)
        return target

    def _remove_template_rows(self, ws) -> None:
        # Header row is row 6. Example/template row is row 7 in our generated sheets.
        if ws.max_row >= 7:
            row_values = [str(ws.cell(7, col).value or "") for col in range(1, ws.max_column + 1)]
            joined = " ".join(row_values).lower()
            if "ej granskad" in joined and (
                "standardflik" in joined
                or "edp-data får aldrig blandas" in joined
                or "granska manuellt" in joined
            ):
                ws.delete_rows(7, 1)
