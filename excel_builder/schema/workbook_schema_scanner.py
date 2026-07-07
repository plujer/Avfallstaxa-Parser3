"""Scan a workbook structure without modifying it."""

from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from excel_builder.models.workbook_schema_models import HeaderCandidate, SheetSchema, WorkbookSchema

class WorkbookSchemaScanner:
    MAX_HEADER_SCAN_ROWS = 30
    MAX_HEADER_SCAN_COLS = 80

    KEY_HEADER_TERMS = {
        "paragraf", "paragrafnamn", "taxapunkt", "variant", "enhet",
        "taxakod", "föreslagen taxa", "foreslagen taxa",
        "strtaxekod", "strtaxebenamning", "strfaktor",
        "strtaxedelavser", "strformel",
    }

    def scan(self, workbook_path: str | Path) -> WorkbookSchema:
        source = Path(workbook_path)
        schema = WorkbookSchema(workbook_path=str(source))
        if not source.exists():
            schema.warnings.append(f"Arbetsbok saknas: {source}")
            return schema

        wb = load_workbook(source, data_only=False, read_only=False)
        schema.defined_names = sorted([name for name in wb.defined_names])

        for ws in wb.worksheets:
            sheet = SheetSchema(
                name=ws.title,
                max_row=ws.max_row,
                max_column=ws.max_column,
                visible_state=ws.sheet_state,
                freeze_panes=str(ws.freeze_panes or ""),
                tables=sorted(list(ws.tables.keys())),
                merged_ranges=[str(item) for item in ws.merged_cells.ranges],
                data_validations=len(ws.data_validations.dataValidation),
                auto_filter_ref=str(ws.auto_filter.ref or ""),
                formula_count=self._formula_count(ws),
                hidden_columns=self._hidden_columns(ws),
                hidden_rows=self._hidden_rows(ws),
            )
            sheet.header_candidates = self._header_candidates(ws)
            if sheet.header_candidates:
                best = sorted(sheet.header_candidates, key=lambda item: item.score, reverse=True)[0]
                sheet.detected_header_row = best.row_number
                sheet.detected_headers = best.values
            schema.sheets.append(sheet)
        return schema

    def _header_candidates(self, ws) -> list[HeaderCandidate]:
        candidates = []
        for row_idx in range(1, min(ws.max_row, self.MAX_HEADER_SCAN_ROWS) + 1):
            values = [self._value(ws.cell(row_idx, col_idx).value)
                      for col_idx in range(1, min(ws.max_column, self.MAX_HEADER_SCAN_COLS) + 1)]
            non_empty = [value for value in values if value]
            if len(non_empty) < 2:
                continue
            score = self._header_score(non_empty)
            if score > 0:
                candidates.append(HeaderCandidate(row_idx, len(non_empty), score, values[:max(10, len(non_empty))]))
        return sorted(candidates, key=lambda item: item.score, reverse=True)

    def _header_score(self, values: list[str]) -> float:
        normalized = [self._norm(value) for value in values]
        joined = " | ".join(normalized)
        score = min(len(values) / 20, 1.0) * 0.25
        matches = sum(1 for term in self.KEY_HEADER_TERMS if term in joined)
        score += min(matches / 6, 1.0) * 0.65
        text_like = sum(1 for value in normalized if any(ch.isalpha() for ch in value))
        score += min(text_like / max(len(values), 1), 1.0) * 0.10
        return round(score, 4)

    def _formula_count(self, ws) -> int:
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    count += 1
        return count

    def _hidden_columns(self, ws) -> list[str]:
        return [get_column_letter(idx) for idx in range(1, ws.max_column + 1)
                if ws.column_dimensions[get_column_letter(idx)].hidden]

    def _hidden_rows(self, ws) -> list[int]:
        return [idx for idx in range(1, ws.max_row + 1) if ws.row_dimensions[idx].hidden]

    def _value(self, value) -> str:
        return "" if value is None else str(value).strip()

    def _norm(self, value: str) -> str:
        return " ".join(value.replace("\\xa0", " ").lower().strip().split())
