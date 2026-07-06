"""Write Tax Knowledge features to generated workbooks.

This writer only writes to a generated workbook copy. It must not modify the
versioned master template directly.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_builder.models import TaxKnowledgeReport


class KnowledgeWorkbookWriter:
    SHEET_NAME = "Tax_Knowledge"

    HEADERS = [
        "Paragraf",
        "Taxapunkt",
        "Variant",
        "Enhet",
        "Paragrafgrupp",
        "Kategori",
        "Avfallstyp",
        "Enhetstyp",
        "Behållarvolym liter",
        "Faktorhint",
        "Confidence",
        "Keywords",
        "Notes",
    ]

    def write(self, workbook_path: str | Path, report: TaxKnowledgeReport) -> Path:
        target = Path(workbook_path)
        wb = load_workbook(target)

        if self.SHEET_NAME in wb.sheetnames:
            wb.remove(wb[self.SHEET_NAME])

        ws = wb.create_sheet(self.SHEET_NAME)

        ws.append(["Tax Knowledge"])
        ws.append(["Syfte", "Strukturerad kunskap från Word/parsern. Används som stöd för kommande regelbaserad matchning."])
        ws.append(["Viktigt", "Denna flik ändrar inte Taxa_från_edp och sätter inga taxekoder."])
        ws.append([])
        ws.append(self.HEADERS)

        for item in report.features:
            ws.append([
                item.parser_row.section,
                item.parser_row.tax_point,
                item.parser_row.variant,
                item.parser_row.unit,
                item.section_group,
                item.category,
                item.waste_type,
                item.unit_type,
                item.container_volume_liter,
                item.factor_hint,
                f"{item.confidence:.2f}",
                ", ".join(item.keywords),
                " | ".join(item.notes),
            ])

        self._style(ws)
        wb.save(target)
        return target

    def _style(self, ws) -> None:
        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        note_fill = PatternFill("solid", fgColor="FFF2CC")

        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill = title_fill

        for row_idx in [2, 3]:
            ws.cell(row_idx, 1).font = Font(bold=True)
            ws.cell(row_idx, 1).fill = note_fill
            ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True)

        header_row = 5
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)

        widths = {
            "A": 12,
            "B": 52,
            "C": 28,
            "D": 16,
            "E": 16,
            "F": 26,
            "G": 26,
            "H": 18,
            "I": 20,
            "J": 18,
            "K": 12,
            "L": 60,
            "M": 60,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A6"

        if ws.max_row >= header_row + 1:
            table = Table(displayName="TaxKnowledgeTable", ref=f"A{header_row}:M{ws.max_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
