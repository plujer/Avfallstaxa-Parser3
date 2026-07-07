"""Read the master workbook as a global rule repository.

Priority order:
1. Taxepunkter rows with existing tax codes become high-priority rules.
2. Taxa_från_edp rows become EDP reference rules.
3. Documentation/reference sheets become lower-priority documentation rules.

This reader does not modify the master workbook.
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from excel_builder.knowledge import TaxKnowledgeExtractor
from excel_builder.matching import MatchNormalizer
from excel_builder.models import MasterRule, ParserTaxRow, RuleRepository
from excel_builder.rules.dynamic_taxepunkter_reader import DynamicTaxepunkterReader


class MasterRuleRepositoryReader:
    TAXEPUNKTER_SHEET = "Taxepunkter"
    EDP_SHEET = "Taxa_från_edp"

    def __init__(self) -> None:
        self.normalizer = MatchNormalizer()
        self.knowledge_extractor = TaxKnowledgeExtractor()
        self.dynamic_taxepunkter_reader = DynamicTaxepunkterReader()

    def read(self, workbook_path: str | Path) -> RuleRepository:
        source = Path(workbook_path)
        repo = RuleRepository(source_workbook=str(source))

        if not source.exists():
            repo.warnings.append(f"Masterarbetsbok saknas: {source}")
            return repo

        wb = load_workbook(source, data_only=True, read_only=False)

        if self.TAXEPUNKTER_SHEET in wb.sheetnames:
            rules, warnings = self.dynamic_taxepunkter_reader.read(source)
            repo.rules.extend(rules)
            repo.warnings.extend(warnings)
        else:
            repo.warnings.append("Fliken Taxepunkter saknas i masterarbetsboken.")

        if self.EDP_SHEET in wb.sheetnames:
            self._read_edp(wb[self.EDP_SHEET], repo)
        else:
            repo.warnings.append("Fliken Taxa_från_edp saknas i masterarbetsboken.")

        self._read_documentation_sheets(wb, repo)

        return repo

    def _read_edp(self, ws, repo: RuleRepository) -> None:
        header_row = self._find_header_row(ws, ["strtaxekod", "strtaxebenamning"])
        if header_row is None:
            repo.warnings.append("Taxa_från_edp saknar EDP-header.")
            return

        headers = self._headers(ws, header_row=header_row)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            tax_code = self._cell(ws, row_idx, headers, ["strtaxekod", "taxakod"])
            name = self._cell(ws, row_idx, headers, ["strtaxebenamning", "taxebenamning", "benämning"])
            factor = self._cell(ws, row_idx, headers, ["strfaktor", "faktor"])
            tax_part = self._cell(ws, row_idx, headers, ["strtaxedelavser", "taxedel"])
            formula = self._cell(ws, row_idx, headers, ["strformel", "formel"])

            if not any([tax_code, name, factor, tax_part, formula]):
                continue

            feature = self._feature("", name, "", "")

            repo.rules.append(
                MasterRule(
                    source_sheet=ws.title,
                    row_number=row_idx,
                    rule_type="EDP",
                    priority=5,
                    tax_point=name,
                    category=feature.category,
                    waste_type=feature.waste_type,
                    unit_type=feature.unit_type,
                    factor_hint=factor or feature.factor_hint,
                    tax_code=tax_code,
                    formula=formula,
                    tax_part=tax_part,
                    source_text=" | ".join([tax_code, name, factor, tax_part, formula]),
                    confidence=1.0,
                )
            )

    def _read_documentation_sheets(self, wb, repo: RuleRepository) -> None:
        skip = {self.TAXEPUNKTER_SHEET, self.EDP_SHEET}
        for ws in wb.worksheets:
            if ws.title in skip:
                continue
            title_norm = self.normalizer.normalize(ws.title)
            if not any(token in title_norm for token in ["dokumentation", "regel", "standard", "referens", "taxa"]):
                continue

            for row_idx in range(1, min(ws.max_row, 250) + 1):
                values = [self._value(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 25) + 1)]
                text = " | ".join([value for value in values if value])
                if len(text) < 8:
                    continue

                repo.rules.append(
                    MasterRule(
                        source_sheet=ws.title,
                        row_number=row_idx,
                        rule_type="DOCUMENTATION",
                        priority=80,
                        source_text=text,
                        confidence=0.50,
                    )
                )

    def _feature(self, section: str, tax_point: str, variant: str, unit: str):
        rows = [ParserTaxRow(section=section, tax_point=tax_point, variant=variant, unit=unit)]
        return self.knowledge_extractor.extract(rows).features[0]

    def _headers(self, ws, header_row: int = 1) -> dict[str, int]:
        headers = {}
        for col in range(1, ws.max_column + 1):
            value = self.normalizer.normalize(str(ws.cell(header_row, col).value or ""))
            if value:
                headers[value] = col
        return headers

    def _find_header_row(self, ws, required: list[str]) -> int | None:
        for row_idx in range(1, min(ws.max_row, 30) + 1):
            values = [self.normalizer.normalize(str(ws.cell(row_idx, col).value or "")) for col in range(1, min(ws.max_column, 40) + 1)]
            joined = " | ".join(values)
            if all(item in joined for item in required):
                return row_idx
        return None

    def _cell(self, ws, row_idx: int, headers: dict[str, int], aliases: list[str]) -> str:
        for alias in aliases:
            alias_norm = self.normalizer.normalize(alias)
            for header, col in headers.items():
                if alias_norm == header or alias_norm in header:
                    return self._value(ws.cell(row_idx, col).value)
        return ""

    def _value(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()
