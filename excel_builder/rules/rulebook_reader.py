"""Read the documented rulebook from Arbets-Excel.

This module is read-only. It does not modify the workbook.
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from excel_builder.models import RuleEntry, Rulebook


class RulebookReader:
    """Extract rules from the workbook documentation sheets."""

    DEFAULT_RULE_SHEETS = [
        "01_Projektöversikt",
        "02_Projektregler",
        "03_Arbetsflöde",
        "04_Källdokument",
        "06_Prompt_ChatGPT",
        "Projektstatus",
        "Dokumentation_Taxepunkter",
        "Dokumentation_Taxa_Saknas",
        "Dokumentation_Taxa_från_edp",
        "Taxekod_Tolkning",
        "Beslutslogg",
        "Färgstandard",
        "Revision_Etapp1",
        "Revision_Formler_DV",
    ]

    def read(self, workbook_path: str | Path = "data/master_templates/ArbetsExcel_Template_v1.0.xlsx") -> Rulebook:
        source = Path(workbook_path)
        rulebook = Rulebook()

        if not source.exists():
            rulebook.warnings.append(f"Arbets-Excel saknas: {source}")
            return rulebook

        wb = load_workbook(source, data_only=True, read_only=False)

        for sheet_name in self.DEFAULT_RULE_SHEETS:
            if sheet_name not in wb.sheetnames:
                rulebook.warnings.append(f"Regelflik saknas: {sheet_name}")
                continue

            ws = wb[sheet_name]
            for row_idx in range(1, ws.max_row + 1):
                values = [
                    self._clean(ws.cell(row_idx, col_idx).value)
                    for col_idx in range(1, min(ws.max_column, 8) + 1)
                ]
                if not any(values):
                    continue

                key = values[0]
                status = values[1] if len(values) > 1 else ""
                text = " | ".join(value for value in values if value)

                rulebook.entries.append(
                    RuleEntry(
                        source_sheet=sheet_name,
                        row_number=row_idx,
                        key=key,
                        status=status,
                        text=text,
                    )
                )

        return rulebook

    def _clean(self, value) -> str:
        return " ".join(str(value or "").replace("\xa0", " ").split())
