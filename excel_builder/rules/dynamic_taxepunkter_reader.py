"""Dynamic Taxepunkter reader using WorkbookSchemaScanner.

This reader fixes the central issue found in Block 24:
Taxepunkter has headers on row 5 in the real master workbook, not row 1.

The reader detects the header row via WorkbookSchemaScanner and then maps columns
by aliases. It only reads the workbook and never modifies the master.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from excel_builder.knowledge import TaxKnowledgeExtractor
from excel_builder.matching import MatchNormalizer
from excel_builder.models import MasterRule, ParserTaxRow
from excel_builder.schema import WorkbookSchemaScanner


class DynamicTaxepunkterReader:
    SHEET_NAME = "Taxepunkter"

    COLUMN_ALIASES = {
        "section": ["paragraf", "section"],
        "paragraph_name": ["paragrafnamn", "paragraph name"],
        "tax_point": ["taxapunkt", "taxepunkt", "namn", "name"],
        "variant": ["variant"],
        "unit": ["enhet", "unit"],
        "tax_code": ["taxakod", "taxa kod", "edp taxekod", "edp_taxekod", "strtaxekod"],
        "proposed_tax": ["föreslagen taxa", "foreslagen taxa"],
        "formula": ["formel", "strformel"],
        "tax_part": ["taxedel", "strtaxedelavser", "taxedel avser"],
        "factor": ["faktor", "strfaktor"],
        "comment": ["kommentar", "notering", "beslutskommentar"],
        "decision_status": ["beslutsstatus", "status"],
    }

    def __init__(self) -> None:
        self.normalizer = MatchNormalizer()
        self.knowledge_extractor = TaxKnowledgeExtractor()

    def read(self, workbook_path: str | Path) -> tuple[list[MasterRule], list[str]]:
        source = Path(workbook_path)
        warnings: list[str] = []

        if not source.exists():
            return [], [f"Arbetsbok saknas: {source}"]

        schema = WorkbookSchemaScanner().scan(source)
        sheet_schema = schema.sheet(self.SHEET_NAME)
        if sheet_schema is None:
            return [], ["Fliken Taxepunkter saknas."]

        header_row = sheet_schema.detected_header_row
        if not header_row:
            return [], ["Kunde inte hitta rubrikrad i Taxepunkter."]

        wb = load_workbook(source, data_only=True, read_only=False)
        ws = wb[self.SHEET_NAME]

        headers = self._headers(ws, header_row)
        rules: list[MasterRule] = []

        for row_idx in range(header_row + 1, ws.max_row + 1):
            values = {
                field: self._cell(ws, row_idx, headers, aliases)
                for field, aliases in self.COLUMN_ALIASES.items()
            }

            section = values["section"]
            paragraph_name = values["paragraph_name"]
            tax_point = values["tax_point"]
            variant = values["variant"]
            unit = values["unit"]
            tax_code = values["tax_code"]
            formula = values["formula"]
            tax_part = values["tax_part"]
            factor = values["factor"]
            comment = values["comment"]
            decision_status = values["decision_status"]
            proposed_tax = values["proposed_tax"]

            if not any([section, paragraph_name, tax_point, variant, unit, tax_code, formula, tax_part, factor, comment, decision_status, proposed_tax]):
                continue

            # Skip pure formula/summary rows with no tax identity.
            if not any([section, tax_point, tax_code, proposed_tax]):
                continue

            feature = self._feature(section, tax_point, variant, unit)

            rules.append(
                MasterRule(
                    source_sheet=ws.title,
                    row_number=row_idx,
                    rule_type="TAXEPUNKT",
                    priority=10 if tax_code else 30,
                    section=section,
                    tax_point=tax_point,
                    category=feature.category,
                    waste_type=feature.waste_type,
                    unit_type=feature.unit_type,
                    factor_hint=factor or feature.factor_hint,
                    container_volume_liter=feature.container_volume_liter,
                    tax_code=tax_code,
                    standard_tax_code="",
                    formula=formula,
                    tax_part=tax_part,
                    source_text=" | ".join([
                        section,
                        paragraph_name,
                        tax_point,
                        variant,
                        unit,
                        tax_code,
                        proposed_tax,
                        factor,
                        formula,
                        tax_part,
                        decision_status,
                        comment,
                    ]),
                    confidence=1.0 if tax_code else 0.75,
                )
            )

        if not rules:
            warnings.append(f"Taxepunkter lästes med rubrikrad {header_row}, men inga regler hittades.")

        return rules, warnings

    def _feature(self, section: str, tax_point: str, variant: str, unit: str):
        rows = [ParserTaxRow(section=section, tax_point=tax_point, variant=variant, unit=unit)]
        return self.knowledge_extractor.extract(rows).features[0]

    def _headers(self, ws, header_row: int) -> dict[str, int]:
        headers: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            raw = self._value(ws.cell(header_row, col_idx).value)
            norm = self.normalizer.normalize(raw)
            if norm:
                headers[norm] = col_idx
        return headers

    def _cell(self, ws, row_idx: int, headers: dict[str, int], aliases: list[str]) -> str:
        for alias in aliases:
            alias_norm = self.normalizer.normalize(alias)
            for header, col_idx in headers.items():
                if header == alias_norm or alias_norm in header or header in alias_norm:
                    return self._value(ws.cell(row_idx, col_idx).value)
        return ""

    def _value(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()
